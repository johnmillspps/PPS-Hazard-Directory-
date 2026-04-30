"""
PPS Rail — NWR Hazard Directory, Access Points & Signal Box Contacts
Streamlit Web Application
=========================================================
Auto-loads all hazard directory CSV files from the data subfolder.
Enter an ELR and mileage range — get hazards, access points,
and signal box contacts in one search. Download each as PDF.
"""

import streamlit as st
import pandas as pd
import re
import os
import io
import glob
from datetime import datetime

# PDF generation
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.lib.colors import HexColor, white, black
from reportlab.platypus import Table, TableStyle, Paragraph
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_LEFT, TA_CENTER
from reportlab.platypus.doctemplate import PageTemplate, BaseDocTemplate, Frame
import base64

# ── Page config ──────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="PPS Rail — Hazard Directory",
    page_icon="⚠️",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ── Brand colours ─────────────────────────────────────────────────────────────
COLOURS = {
    'bg':       '#0a0a0a',
    'surface':  '#141414',
    'surface2': '#1e1e1e',
    'border':   '#2a2a2a',
    'red':      '#CC2200',
    'amber':    '#F5A800',
    'green':    '#00A651',
    'nwr_blue': '#003366',
    'text':     '#e0e0e0',
    'muted':    '#888888',
    'white':    '#ffffff',
}

# ── Custom CSS ────────────────────────────────────────────────────────────────
st.markdown(f"""
<style>
  @import url('https://fonts.googleapis.com/css2?family=Barlow+Condensed:wght@400;600;700&family=Barlow:wght@300;400;500&display=swap');

  html, body, [class*="css"] {{
    font-family: 'Barlow', sans-serif;
    background-color: {COLOURS['bg']};
    color: {COLOURS['text']};
  }}

  header[data-testid="stHeader"] {{
    background-color: {COLOURS['bg']} !important;
    color: {COLOURS['bg']} !important;
  }}

  .stDeployButton {{
    display: none !important;
  }}

  section[data-testid="stSidebar"] {{
    background-color: {COLOURS['surface']} !important;
    border-right: 1px solid {COLOURS['border']};
  }}
  section[data-testid="stSidebar"] * {{
    color: {COLOURS['text']} !important;
  }}

  .main .block-container {{
    background-color: {COLOURS['bg']};
    padding-top: 1rem;
  }}

  [data-testid="stAppViewContainer"],
  [data-testid="stAppViewBlockContainer"],
  [data-testid="block-container"] {{
    background-color: {COLOURS['bg']} !important;
  }}

  .stApp {{
    background-color: {COLOURS['bg']} !important;
  }}

  h1, h2, h3, h4 {{
    font-family: 'Barlow Condensed', sans-serif !important;
    letter-spacing: 0.04em;
    color: {COLOURS['white']} !important;
  }}

  .pps-card {{
    background: {COLOURS['surface']};
    border: 1px solid {COLOURS['border']};
    border-radius: 4px;
    padding: 1.2rem 1.4rem;
    margin-bottom: 1rem;
  }}
  .pps-card-green {{ border-left: 3px solid {COLOURS['green']}; }}
  .pps-card-amber {{ border-left: 3px solid {COLOURS['amber']}; }}
  .pps-card-grey {{ border-left: 3px solid {COLOURS['muted']}; }}

  .badge {{
    display: inline-block;
    padding: 2px 10px;
    border-radius: 2px;
    font-size: 0.75rem;
    font-weight: 600;
    letter-spacing: 0.06em;
    text-transform: uppercase;
  }}
  .badge-green {{ background: {COLOURS['green']}22; color: {COLOURS['green']}; border: 1px solid {COLOURS['green']}55; }}
  .badge-amber {{ background: {COLOURS['amber']}22; color: {COLOURS['amber']}; border: 1px solid {COLOURS['amber']}55; }}
  .badge-grey {{ background: {COLOURS['muted']}22; color: {COLOURS['muted']}; border: 1px solid {COLOURS['muted']}55; }}

  .metric-box {{
    background: {COLOURS['surface']};
    border: 1px solid {COLOURS['border']};
    border-radius: 4px;
    padding: 0.8rem 1rem;
    text-align: center;
  }}
  .metric-num {{
    font-family: 'Barlow Condensed', sans-serif;
    font-size: 2rem;
    font-weight: 700;
    line-height: 1;
  }}
  .metric-lbl {{
    font-size: 0.72rem;
    color: {COLOURS['muted']};
    text-transform: uppercase;
    letter-spacing: 0.08em;
    margin-top: 0.2rem;
  }}

  .pps-divider {{
    border: none;
    border-top: 1px solid {COLOURS['border']};
    margin: 1rem 0;
  }}

  .section-header {{
    font-family: 'Barlow Condensed', sans-serif;
    font-size: 1.1rem;
    font-weight: 700;
    color: {COLOURS['white']};
    letter-spacing: 0.04em;
    padding: 0.6rem 1rem;
    background: {COLOURS['surface']};
    border-left: 3px solid {COLOURS['nwr_blue']};
    margin: 1.5rem 0 0.8rem 0;
  }}

  .stTextInput input {{
    background-color: {COLOURS['surface2']} !important;
    color: {COLOURS['white']} !important;
    border: 1px solid {COLOURS['border']} !important;
    border-radius: 3px !important;
  }}
  .stTextInput input::placeholder {{
    color: {COLOURS['muted']} !important;
    opacity: 1 !important;
  }}
  .stTextInput label {{
    color: {COLOURS['white']} !important;
    font-size: 0.85rem !important;
    font-weight: 500 !important;
    letter-spacing: 0.04em !important;
  }}

  .stButton > button {{
    background-color: {COLOURS['nwr_blue']} !important;
    color: white !important;
    border: none !important;
    border-radius: 3px !important;
    font-family: 'Barlow Condensed', sans-serif !important;
    font-size: 1rem !important;
    font-weight: 600 !important;
    letter-spacing: 0.06em !important;
    padding: 0.5rem 2rem !important;
    text-transform: uppercase !important;
  }}
  .stButton > button:hover {{
    background-color: #002244 !important;
  }}

  /* Download buttons — same blue style */
  .stDownloadButton > button {{
    background-color: {COLOURS['nwr_blue']} !important;
    color: white !important;
    border: none !important;
    border-radius: 3px !important;
    font-family: 'Barlow Condensed', sans-serif !important;
    font-size: 0.95rem !important;
    font-weight: 600 !important;
    letter-spacing: 0.06em !important;
    padding: 0.5rem 2rem !important;
    text-transform: uppercase !important;
    width: 100% !important;
  }}
  .stDownloadButton > button:hover {{
    background-color: #002244 !important;
  }}
  .stDownloadButton > button span,
  .stDownloadButton > button div,
  .stDownloadButton > button p {{
    color: white !important;
  }}

  label, .stMarkdown p {{
    color: {COLOURS['text']} !important;
  }}

  p, span, div, li {{
    color: {COLOURS['text']};
  }}

  .stDataFrame {{
    color: {COLOURS['white']} !important;
  }}

  /* ── Selectbox: white box, black text ── */
  .stSelectbox [data-baseweb="select"] {{
    background-color: #ffffff !important;
  }}
  .stSelectbox [data-baseweb="select"] * {{
    color: #000000 !important;
  }}
  .stSelectbox [data-baseweb="select"] svg {{
    fill: #000000 !important;
  }}

  /* ── Dropdown menu / popover: BLACK text on WHITE ── */
  [data-baseweb="popover"],
  [data-baseweb="popover"] > div,
  [data-baseweb="menu"],
  [data-baseweb="menu"] ul,
  [data-baseweb="menu"] li,
  [data-baseweb="list"],
  [data-baseweb="list"] li {{
    background-color: #ffffff !important;
    color: #000000 !important;
  }}
  /* Hover state for menu items */
  [data-baseweb="menu"] li:hover,
  [data-baseweb="list"] li:hover {{
    background-color: #e8e8e8 !important;
    color: #000000 !important;
  }}
  /* Highlighted / focused option */
  [data-baseweb="menu"] li[aria-selected="true"],
  [data-baseweb="menu"] li[data-highlighted="true"] {{
    background-color: #d0d0d0 !important;
    color: #000000 !important;
  }}
  /* Catch-all for any nested spans/divs inside menus */
  [data-baseweb="popover"] span,
  [data-baseweb="popover"] div,
  [data-baseweb="menu"] span,
  [data-baseweb="menu"] div {{
    color: #000000 !important;
  }}

  /* ── MultiSelect ── */
  .stMultiSelect [data-baseweb="select"] {{
    background-color: #ffffff !important;
  }}
  .stMultiSelect [data-baseweb="select"] * {{
    color: #000000 !important;
  }}
  /* Green tags (NOT red — red means danger on the railway!) */
  .stMultiSelect [data-baseweb="tag"],
  .stMultiSelect span[data-baseweb="tag"] {{
    background-color: {COLOURS['green']} !important;
  }}
  .stMultiSelect [data-baseweb="tag"] *,
  .stMultiSelect span[data-baseweb="tag"] * {{
    color: #ffffff !important;
  }}

  /* ── TextArea ── */
  .stTextArea textarea {{
    background-color: #ffffff !important;
    color: #000000 !important;
    border: 1px solid #cccccc !important;
  }}

  /* ── TextInput (SWP tab overrides the earlier dark-theme style) ── */
  .stTextInput input {{
    background-color: #ffffff !important;
    color: #000000 !important;
    border: 1px solid #cccccc !important;
  }}
</style>
""", unsafe_allow_html=True)


# ── Mileage helpers ──────────────────────────────────────────────────────────
def mileage_to_decimal(s):
    """Convert '182m 10ch' to decimal miles (182.0220) for hazard CSV comparison."""
    if not s:
        return None
    m = re.search(r'(\d+)\s*m\s*([\d.]+)\s*c(?:h)?', str(s), re.IGNORECASE)
    if m:
        miles = int(m.group(1))
        chains = float(m.group(2))
        yards = chains * 22
        return miles + yards / 10000
    return None


def decimal_to_miles_chains(dec):
    """Convert decimal miles.yards (182.0259) to 'Xm Ych' display format."""
    if dec is None or dec == '':
        return ''
    try:
        dec = float(dec)
        miles = int(dec)
        yards = round((abs(dec) - abs(miles)) * 10000)
        chains = round(yards / 22)
        return f"{miles}m {chains:02d}ch"
    except (ValueError, TypeError):
        return str(dec)


# ── Auto-load hazard CSVs ────────────────────────────────────────────────────
@st.cache_data
def load_all_hazard_csvs():
    """Load all CSV files from the data subfolder and combine into one DataFrame."""
    data_dir = os.path.join(os.path.dirname(__file__), 'data')
    if not os.path.exists(data_dir):
        return None, [], 0

    csv_files = glob.glob(os.path.join(data_dir, '*.csv'))
    if not csv_files:
        return None, [], 0

    dfs = []
    loaded_files = []
    for f in sorted(csv_files):
        try:
            df = pd.read_csv(f, quotechar='"', encoding='utf-8')
            if 'ELR' in df.columns:
                dfs.append(df)
                loaded_files.append(os.path.basename(f))
        except Exception:
            pass

    if not dfs:
        return None, [], 0

    combined = pd.concat(dfs, ignore_index=True)
    if 'Hazard ID' in combined.columns:
        combined = combined.drop_duplicates(subset=['Hazard ID'])

    return combined, loaded_files, len(csv_files)


# ── Auto-load signal box contacts ────────────────────────────────────────────
@st.cache_data
def load_signal_box_contacts():
    """Load signal box contacts CSV from the data subfolder."""
    data_dir = os.path.join(os.path.dirname(__file__), 'data')
    csv_path = os.path.join(data_dir, 'signal_box_contacts.csv')
    if not os.path.exists(csv_path):
        return None
    try:
        df = pd.read_csv(csv_path, encoding='utf-8-sig')
        return df
    except Exception:
        return None


# ── Auto-load A&E departments ────────────────────────────────────────────────
@st.cache_data
def load_ae_departments():
    """Load Type 1 A&E departments CSV from the data subfolder."""
    data_dir = os.path.join(os.path.dirname(__file__), 'data')
    csv_path = os.path.join(data_dir, 'ae_departments_type1.csv')
    if not os.path.exists(csv_path):
        return None
    try:
        df = pd.read_csv(csv_path, encoding='utf-8-sig')
        return df
    except Exception:
        return None


# ── Auto-load access points with coordinates ─────────────────────────────────
@st.cache_data
def load_access_points_coords():
    """Load access points with lat/lon from the data subfolder."""
    data_dir = os.path.join(os.path.dirname(__file__), 'data')
    csv_path = os.path.join(data_dir, 'access_points_with_coords.csv')
    if not os.path.exists(csv_path):
        return None
    try:
        df = pd.read_csv(csv_path, encoding='utf-8-sig')
        return df
    except Exception:
        return None


# ── Auto-load signal box areas ───────────────────────────────────────────────
@st.cache_data
def load_signal_box_areas():
    """Load signal box areas CSV from the data subfolder."""
    data_dir = os.path.join(os.path.dirname(__file__), 'data')
    csv_path = os.path.join(data_dir, 'signal_box_areas.csv')
    if not os.path.exists(csv_path):
        return None
    try:
        df = pd.read_csv(csv_path, encoding='utf-8-sig')
        return df
    except Exception:
        return None


# ── Auto-load line names ─────────────────────────────────────────────────────
@st.cache_data
def load_line_names():
    """Load line names CSV from the data subfolder."""
    data_dir = os.path.join(os.path.dirname(__file__), 'data')
    csv_path = os.path.join(data_dir, 'line_names.csv')
    if not os.path.exists(csv_path):
        return None
    try:
        df = pd.read_csv(csv_path, encoding='utf-8-sig')
        return df
    except Exception:
        return None


@st.cache_data
def load_signal_diagram_index():
    """Load signal diagram index CSV from the data subfolder."""
    data_dir = os.path.join(os.path.dirname(__file__), 'data')
    csv_path = os.path.join(data_dir, 'signal_diagram_index.csv')
    if not os.path.exists(csv_path):
        return None
    try:
        df = pd.read_csv(csv_path, encoding='utf-8-sig')
        return df
    except Exception:
        return None


@st.cache_data
def load_signal_ref_lookup():
    """Load signal ref to diagram page lookup CSV."""
    data_dir = os.path.join(os.path.dirname(__file__), 'data')
    csv_path = os.path.join(data_dir, 'signal_ref_lookup.csv')
    if not os.path.exists(csv_path):
        return None
    try:
        df = pd.read_csv(csv_path, encoding='utf-8-sig')
        return df
    except Exception:
        return None


def find_nearest_ae(lat, lon, ae_df, n=3):
    """Find the n nearest Type 1 A&E departments to a given lat/lon using haversine."""
    import math
    R = 3959  # miles

    def haversine(lat1, lon1, lat2, lon2):
        dlat = math.radians(lat2 - lat1)
        dlon = math.radians(lon2 - lon1)
        a = (math.sin(dlat / 2) ** 2 +
             math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) *
             math.sin(dlon / 2) ** 2)
        return R * 2 * math.asin(math.sqrt(a))

    results = []
    for _, row in ae_df.iterrows():
        try:
            d = haversine(lat, lon, float(row['latitude']), float(row['longitude']))
            results.append({
                'Hospital': row['hospital_name'],
                'Trust': row['trust_name'],
                'Address': row['address'],
                'Postcode': row['postcode'],
                'Distance (miles)': round(d, 1),
            })
        except (ValueError, TypeError):
            pass
    results.sort(key=lambda x: x['Distance (miles)'])
    return results[:n]


def find_worksite_coords(elr_from, elr_to, from_dec, to_dec, ap_coords_df):
    """Find the best lat/lon for a worksite by matching access points with coordinates."""
    if ap_coords_df is None or ap_coords_df.empty:
        return None, None, None

    mid_dec = (from_dec + to_dec) / 2

    # Search both ELRs
    elrs = [elr_from]
    if elr_to and elr_to != elr_from:
        elrs.append(elr_to)

    best_dist = float('inf')
    best_lat, best_lon, best_name = None, None, None

    for elr_q in elrs:
        matches = ap_coords_df[ap_coords_df['elr'] == elr_q].copy()
        if matches.empty:
            continue
        matches['dist'] = (matches['mileage_decimal'] - mid_dec).abs()
        closest = matches.nsmallest(1, 'dist').iloc[0]
        if closest['dist'] < best_dist:
            best_dist = closest['dist']
            best_lat = closest['latitude']
            best_lon = closest['longitude']
            best_name = f"{closest['name']} ({closest['mileage_ch']})"

    if best_lat and best_lon:
        return best_lat, best_lon, best_name
    return None, None, None


def enrich_access_points_with_coords(access_pts_df, ap_coords_df):
    """Add lat/lon and Google Maps link to access points by matching ELR + mileage."""
    if ap_coords_df is None or ap_coords_df.empty or access_pts_df.empty:
        return access_pts_df

    lats, lons, links = [], [], []

    for _, row in access_pts_df.iterrows():
        elr = row.get('ELR', '')
        mil_dec = pd.to_numeric(row.get('Mileage  From', None), errors='coerce')

        if pd.isna(mil_dec) or not elr:
            lats.append('')
            lons.append('')
            links.append('')
            continue

        matches = ap_coords_df[ap_coords_df['elr'] == elr].copy()
        if matches.empty:
            lats.append('')
            lons.append('')
            links.append('')
            continue

        matches['dist'] = (matches['mileage_decimal'] - mil_dec).abs()
        closest = matches.nsmallest(1, 'dist').iloc[0]

        # Only match if within ~0.5 decimal miles (about 0.5 miles)
        if closest['dist'] < 0.5:
            lat = closest['latitude']
            lon = closest['longitude']
            lats.append(f"{lat:.5f}")
            lons.append(f"{lon:.5f}")
            links.append(f"https://maps.google.com/?q={lat:.4f},{lon:.4f}")
        else:
            lats.append('')
            lons.append('')
            links.append('')

    result = access_pts_df.copy()
    result['Lat'] = lats
    result['Lon'] = lons
    result['Google Maps'] = links
    return result


def format_phone_numbers(raw_phone):
    """Format phone numbers that may be concatenated into separate lines.

    Splits strings like '033 085 41095 (emergency only) 033 085 41096 01270 255 582'
    into individual numbers, each on its own line.
    """
    if not raw_phone or raw_phone == 'nan':
        return ''
    raw_phone = str(raw_phone).strip()
    if not raw_phone:
        return ''

    # Find all UK phone numbers: start with 0, then digits/spaces totalling 10-11 digits
    # Optionally followed by a parenthetical note like '(emergency only)'
    numbers = re.findall(
        r'(0(?:\d[\s]?){9,10}(?:\s*\([^)]+\))?)', raw_phone
    )
    if len(numbers) <= 1:
        return raw_phone
    return '<br/>'.join(n.strip() for n in numbers)


def find_signal_boxes_for_mileage(elr_from, elr_to, from_ch, to_ch, sba_df, signalbox_df):
    """Find signal boxes covering a mileage range and match with contact details."""
    if sba_df is None or sba_df.empty:
        return []

    elrs = [elr_from]
    if elr_to and elr_to != elr_from:
        elrs.append(elr_to)

    found_boxes = []
    seen_names = set()

    for elr_q in elrs:
        matches = sba_df[
            (sba_df['elr'] == elr_q) &
            (sba_df['mileage_from_ch'] <= to_ch) &
            (sba_df['mileage_to_ch'] >= from_ch)
        ]
        # Fallback: if no mileage overlap found, show all signal boxes on this ELR
        if matches.empty:
            matches = sba_df[sba_df['elr'] == elr_q]
        for _, row in matches.iterrows():
            box_name = str(row['signal_box_name']).strip()
            if box_name in seen_names or len(box_name) < 3:
                continue
            seen_names.add(box_name)

            phone = ''
            prefix = str(row.get('signal_prefix', '')).strip()
            if prefix == 'nan':
                prefix = ''
            eco_name = str(row.get('eco_name', '')).strip()
            if eco_name == 'nan':
                eco_name = ''
            eco_type = str(row.get('eco_type', '')).strip()
            if eco_type == 'nan':
                eco_type = ''

            if signalbox_df is not None and not signalbox_df.empty and box_name:
                first_word = box_name.split()[0] if box_name else ''
                contact_match = signalbox_df[
                    signalbox_df['Signal Box'].str.contains(first_word, case=False, na=False)
                ]
                if not contact_match.empty:
                    phone = str(contact_match.iloc[0].get('External Telephone', '')).strip()
                if (not phone or phone == 'nan') and prefix:
                    prefix_match = signalbox_df[
                        signalbox_df['Signal Prefix'].str.upper().eq(prefix.upper())
                    ]
                    if not prefix_match.empty:
                        phone = str(prefix_match.iloc[0].get('External Telephone', '')).strip()
            if phone == 'nan':
                phone = ''

            eco_str = f"{eco_name.rstrip('. ')} {eco_type}".strip()

            # Look up ECR phone number from contacts
            eco_phone = ''
            if eco_name and signalbox_df is not None and not signalbox_df.empty:
                # Search for ECR/ECO entry matching the eco_name
                eco_search = eco_name.split()[0] if eco_name else ''
                if eco_search:
                    ecr_match = signalbox_df[
                        signalbox_df['Signal Box'].str.contains(eco_search, case=False, na=False) &
                        signalbox_df['Signal Box'].str.contains('ECR|ECO|Electrical', case=False, na=False, regex=True)
                    ]
                    if not ecr_match.empty:
                        eco_phone = str(ecr_match.iloc[0].get('External Telephone', '')).strip()
                    if not eco_phone or eco_phone == 'nan':
                        # Try matching just the location name + ECR
                        ecr_match2 = signalbox_df[
                            signalbox_df['Signal Box'].str.contains(f'{eco_search}.*ECR', case=False, na=False, regex=True)
                        ]
                        if not ecr_match2.empty:
                            eco_phone = str(ecr_match2.iloc[0].get('External Telephone', '')).strip()
                if eco_phone == 'nan':
                    eco_phone = ''

            found_boxes.append({
                'Signal Box': box_name,
                'Prefix': prefix,
                'Phone': phone,
                'ECO': eco_str,
                'ECO Phone': eco_phone,
                'ELR': elr_q,
                'Coverage': f"{row.get('mileage_from_raw', '')} to {row.get('mileage_to_raw', '')}",
            })

    return found_boxes


def find_line_names_for_mileage(elr_from, elr_to, from_ch, to_ch, ln_df):
    """Find line names covering a mileage range."""
    if ln_df is None or ln_df.empty:
        return []

    elrs = [elr_from]
    if elr_to and elr_to != elr_from:
        elrs.append(elr_to)

    found_lines = []
    seen = set()

    for elr_q in elrs:
        matches = ln_df[
            (ln_df['elr'] == elr_q) &
            (ln_df['mileage_from_ch'] <= to_ch) &
            (ln_df['mileage_to_ch'] >= from_ch)
        ]
        # Fallback: if no mileage overlap found, show all line names on this ELR
        if matches.empty:
            matches = ln_df[ln_df['elr'] == elr_q]
        for _, row in matches.iterrows():
            abbr = str(row['abbreviation']).strip()
            full = str(row['full_name']).strip()
            key = f"{abbr}_{full}"
            if key not in seen:
                seen.add(key)
                found_lines.append({
                    'Abbreviation': abbr,
                    'Line Name': full,
                })

    return found_lines


# ── Filter helpers ───────────────────────────────────────────────────────────
def filter_access_points(df):
    """Filter DataFrame to only access point rows."""
    if 'Hazard Description' not in df.columns:
        return df.iloc[0:0]
    mask = df['Hazard Description'].str.contains('Access Point', case=False, na=False)
    return df[mask].copy()


def filter_hazards_only(df):
    """Filter DataFrame to exclude access point rows (hazards only)."""
    if 'Hazard Description' not in df.columns:
        return df
    mask = ~df['Hazard Description'].str.contains('Access Point', case=False, na=False)
    return df[mask].copy()


def query_by_elr_mileage(df, elr_from, elr_to, from_dec, to_dec):
    """Query a DataFrame by ELR(s) and mileage range with overlap logic."""
    elrs_to_query = [elr_from]
    if elr_to and elr_to != elr_from:
        elrs_to_query.append(elr_to)

    frames = []
    for elr_q in elrs_to_query:
        elr_data = df[df['ELR'] == elr_q].copy()
        if not elr_data.empty:
            elr_data['mile_from'] = pd.to_numeric(
                elr_data['Mileage  From'], errors='coerce')
            elr_data['mile_to'] = pd.to_numeric(
                elr_data['Mileage To'], errors='coerce')

            if elr_q == elr_from and elr_q == elr_to:
                filt = elr_data[
                    ~(elr_data['mile_to'] < from_dec) &
                    ~(elr_data['mile_from'] > to_dec)
                ]
            elif elr_q == elr_from:
                filt = elr_data[elr_data['mile_from'] >= from_dec]
            else:
                filt = elr_data[elr_data['mile_to'] <= to_dec]

            frames.append(filt)

    if frames:
        return pd.concat(frames, ignore_index=True)
    return pd.DataFrame()


# ── PDF generation ───────────────────────────────────────────────────────────
NWR_BLUE_PDF = HexColor("#003366")
ALT_ROW_PDF = HexColor("#EEF2F7")
GRID_GREY_PDF = HexColor("#CCCCCC")


class HazardDocTemplate(BaseDocTemplate):
    def __init__(self, filename, **kwargs):
        self.elr = kwargs.pop('elr', '')
        self.issue_date = kwargs.pop('issue_date', '')
        self.from_mil = kwargs.pop('from_mil', '')
        self.to_mil = kwargs.pop('to_mil', '')
        self.pdf_title = kwargs.pop('pdf_title', 'NWR Hazard Directory')
        BaseDocTemplate.__init__(self, filename, **kwargs)


def on_page(canvas, doc):
    canvas.saveState()
    w, h = landscape(A4)
    margins = 10 * mm

    bar_h = 14 * mm
    canvas.setFillColor(NWR_BLUE_PDF)
    canvas.rect(0, h - bar_h, w, bar_h, fill=1, stroke=0)
    canvas.setFillColor(white)
    canvas.setFont('Helvetica-Bold', 10)
    title = f"{doc.pdf_title}  \u2013  ELR: {doc.elr}  \u2013  Issue Date: {doc.issue_date}"
    canvas.drawCentredString(w / 2, h - bar_h + 4 * mm, title)

    canvas.setFillColor(black)
    canvas.setFont('Helvetica-Bold', 8)
    range_text = f"Extract: {doc.from_mil} to {doc.to_mil}"
    canvas.drawCentredString(w / 2, h - bar_h - 4 * mm, range_text)

    canvas.setFillColor(black)
    canvas.setFont('Helvetica', 7)
    canvas.drawString(margins, 6 * mm,
                      f"PPS Rail  |  {doc.issue_date}")
    canvas.drawRightString(w - margins, 6 * mm,
                           f"Page {canvas.getPageNumber()}")
    canvas.restoreState()


def generate_pdf(filtered_df, elr, from_mil, to_mil, col_config, pdf_title='NWR Hazard Directory'):
    """Generate PDF in memory and return bytes."""
    buf = io.BytesIO()
    issue_date = datetime.now().strftime("%d/%m/%Y")

    page_w, page_h = landscape(A4)
    margins = 10 * mm
    usable_w = page_w - 2 * margins

    col_widths = [usable_w * c['width_pct'] for c in col_config]

    header_style = ParagraphStyle(
        'HeaderCell', fontName='Helvetica-Bold', fontSize=7.5,
        textColor=white, leading=9, alignment=TA_LEFT,
    )
    cell_style = ParagraphStyle(
        'DataCell', fontName='Helvetica', fontSize=7,
        textColor=black, leading=8.5, alignment=TA_LEFT,
        wordWrap='CJK',
    )

    link_style = ParagraphStyle(
        'LinkCell', fontName='Helvetica', fontSize=6,
        textColor=HexColor("#0000EE"), leading=7.5, alignment=TA_LEFT,
    )

    doc = HazardDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=margins, rightMargin=margins,
        topMargin=22 * mm, bottomMargin=14 * mm,
        elr=elr, issue_date=issue_date,
        from_mil=from_mil, to_mil=to_mil,
        pdf_title=pdf_title,
    )

    frame = Frame(margins, 14 * mm, usable_w, page_h - 36 * mm, id='main')
    template = PageTemplate(id='main', frames=[frame], onPage=on_page)
    doc.addPageTemplates([template])

    header_row = [Paragraph(c['header'], header_style) for c in col_config]
    table_data = [header_row]

    for _, row in filtered_df.iterrows():
        data_row = []
        for c in col_config:
            val = str(row.get(c['field'], '')).replace('\n', '<br/>')
            if val.startswith('https://'):
                data_row.append(Paragraph(f'<a href="{val}" color="blue"><u>Map Link</u></a>', link_style))
            else:
                data_row.append(Paragraph(val, cell_style))
        table_data.append(data_row)

    table = Table(table_data, colWidths=col_widths, repeatRows=1)

    style_commands = [
        ('BACKGROUND', (0, 0), (-1, 0), NWR_BLUE_PDF),
        ('TEXTCOLOR', (0, 0), (-1, 0), white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 7.5),
        ('ALIGN', (0, 0), (-1, 0), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('GRID', (0, 0), (-1, -1), 0.4, GRID_GREY_PDF),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
    ]

    for i in range(1, len(table_data)):
        if i % 2 == 0:
            style_commands.append(
                ('BACKGROUND', (0, i), (-1, i), ALT_ROW_PDF))

    table.setStyle(TableStyle(style_commands))
    doc.build([table])
    buf.seek(0)
    return buf


# ── Column configs for PDFs ──────────────────────────────────────────────────
HAZARD_COLS = [
    {'header': 'ELR',         'field': 'ELR',                'width_pct': 0.055},
    {'header': 'ELR Name',    'field': 'ELR Name',           'width_pct': 0.18},
    {'header': 'Start',       'field': 'Mileage  From',      'width_pct': 0.055},
    {'header': 'End',         'field': 'Mileage To',         'width_pct': 0.055},
    {'header': 'Description', 'field': 'Hazard Description', 'width_pct': 0.12},
    {'header': 'Local Name',  'field': 'Local Name',         'width_pct': 0.10},
    {'header': 'Track',       'field': 'Track',              'width_pct': 0.07},
    {'header': 'Free Text',   'field': 'Free Text',          'width_pct': 0.365},
]

ACCESS_COLS = [
    {'header': 'ELR',         'field': 'ELR',                'width_pct': 0.06},
    {'header': 'ELR Name',    'field': 'ELR Name',           'width_pct': 0.15},
    {'header': 'Mileage',     'field': 'Mileage  From',      'width_pct': 0.07},
    {'header': 'Type',        'field': 'Hazard Description', 'width_pct': 0.14},
    {'header': 'Local Name',  'field': 'Local Name',         'width_pct': 0.14},
    {'header': 'Track',       'field': 'Track',              'width_pct': 0.08},
    {'header': 'Details',     'field': 'Free Text',          'width_pct': 0.18},
    {'header': 'Google Maps', 'field': 'Google Maps',        'width_pct': 0.18},
]

SIGNALBOX_COLS = [
    {'header': 'Signal Box / Panel / Workstation', 'field': 'Signal Box',         'width_pct': 0.40},
    {'header': 'External Telephone',               'field': 'External Telephone', 'width_pct': 0.25},
    {'header': 'Signal Prefix',                    'field': 'Signal Prefix',      'width_pct': 0.15},
    {'header': 'Route',                            'field': 'Source',             'width_pct': 0.20},
]


# ── APP LAYOUT ───────────────────────────────────────────────────────────────

# Top bar
logo_path = os.path.join(os.path.dirname(__file__), 'PPS-logo-ol.png')
logo_b64 = ''
if os.path.exists(logo_path):
    with open(logo_path, 'rb') as f:
        logo_b64 = base64.b64encode(f.read()).decode()

st.markdown(f"""
<div style="
  display: flex;
  align-items: center;
  gap: 1.5rem;
  padding: 1rem 0 1.2rem 0;
  border-bottom: 1px solid {COLOURS['border']};
  margin-bottom: 1.5rem;
  background: {COLOURS['bg']};
">
  {'<img src="data:image/png;base64,' + logo_b64 + '" style="height:60px;width:auto;border-radius:3px;" />' if logo_b64 else ''}
  <div>
    <div style="font-family: Barlow Condensed, sans-serif; font-size: 1.6rem; font-weight: 700; color: white; letter-spacing: 0.05em;">
      Worksite Intelligence</div>
    <div style="font-size: 0.8rem; color: {COLOURS['muted']}; letter-spacing: 0.1em; text-transform: uppercase;">
      Hazards &bull; Access Points &bull; Signal Boxes &bull; Lines &bull; Nearest A&amp;E &bull; SWP Builder</div>
  </div>
</div>
""", unsafe_allow_html=True)

# ── Load data ────────────────────────────────────────────────────────────────
hazard_df, loaded_files, total_files = load_all_hazard_csvs()
signalbox_df = load_signal_box_contacts()
ae_df = load_ae_departments()
ap_coords_df = load_access_points_coords()
sba_df = load_signal_box_areas()
ln_df = load_line_names()
sd_idx_df = load_signal_diagram_index()
sig_ref_df = load_signal_ref_lookup()

# ── Sidebar ──────────────────────────────────────────────────────────────────
with st.sidebar:
    if hazard_df is not None and not hazard_df.empty:
        n_hazards = len(hazard_df)
        n_elrs = hazard_df['ELR'].nunique()
        n_access = len(filter_access_points(hazard_df))
        n_sb = len(signalbox_df) if signalbox_df is not None else 0
        n_ae = len(ae_df) if ae_df is not None else 0

        st.markdown("### Data Loaded")
        st.markdown(f"""
        <div class="pps-card pps-card-green">
          <span class="badge badge-green">&#10003; {len(loaded_files)} files loaded</span>
        </div>
        """, unsafe_allow_html=True)

        st.markdown(f"""
        <div class="metric-box" style="margin-bottom:0.5rem">
          <div class="metric-num" style="color:{COLOURS['green']}">{n_hazards:,}</div>
          <div class="metric-lbl">Hazards</div>
        </div>
        <div class="metric-box" style="margin-bottom:0.5rem">
          <div class="metric-num" style="color:{COLOURS['amber']}">{n_access:,}</div>
          <div class="metric-lbl">Access Points</div>
        </div>
        <div class="metric-box" style="margin-bottom:0.5rem">
          <div class="metric-num" style="color:{COLOURS['text']}">{n_sb:,}</div>
          <div class="metric-lbl">Signal Boxes</div>
        </div>
        <div class="metric-box" style="margin-bottom:0.5rem">
          <div class="metric-num" style="color:{COLOURS['red']}">{n_ae:,}</div>
          <div class="metric-lbl">A&amp;E Depts</div>
        </div>
        <div class="metric-box">
          <div class="metric-num" style="color:{COLOURS['text']}">{n_elrs:,}</div>
          <div class="metric-lbl">ELRs</div>
        </div>
        """, unsafe_allow_html=True)

        st.markdown('<hr class="pps-divider">', unsafe_allow_html=True)
        st.markdown("### Available ELRs")
        elrs = sorted(hazard_df['ELR'].dropna().unique().tolist())
        for e in elrs[:50]:
            count = len(hazard_df[hazard_df['ELR'] == e])
            st.markdown(f"<small><b>{e}</b> <span style='color:{COLOURS['muted']}'>"
                        f"({count})</span></small>",
                        unsafe_allow_html=True)
        if len(elrs) > 50:
            st.markdown(f"<small style='color:{COLOURS['muted']}'>...and {len(elrs)-50} more</small>",
                        unsafe_allow_html=True)
    else:
        st.markdown("### No Data")
        st.markdown(f"""
        <div class="pps-card pps-card-amber">
          <span class="badge badge-amber">No CSV files found</span>
          <div style="margin-top:0.5rem;font-size:0.85rem;color:{COLOURS['muted']}">
            Place hazard directory CSV files in the <b>data</b> subfolder.
          </div>
        </div>
        """, unsafe_allow_html=True)


# ── Main ─────────────────────────────────────────────────────────────────────
if hazard_df is None or hazard_df.empty:
    st.markdown(f"""
    <div class="pps-card pps-card-grey">
      <span class="badge badge-grey">No data loaded</span>
      <span style="margin-left:1rem;font-size:0.85rem;color:#888">
        Place your hazard directory CSV files in the <b>data</b> subfolder and restart the app.
      </span>
    </div>
    """, unsafe_allow_html=True)
else:
    tab_ws, tab_swp = st.tabs(["🔍  Worksite Intelligence", "📋  SWP Builder"])

    with tab_ws:
        st.markdown("#### Enter your worksite mileage range")

        c1, c2, c3, c4 = st.columns([2, 2, 2, 2])
        with c1:
            elr_from = st.text_input("ELR FROM", placeholder="e.g. CGJ3").upper()
        with c2:
            mil_from = st.text_input("Mileage FROM", placeholder="e.g. 182m 10ch")
        with c3:
            elr_to = st.text_input("ELR TO", placeholder="e.g. CGJ3 (or different)").upper()
        with c4:
            mil_to = st.text_input("Mileage TO", placeholder="e.g. 182m 30ch")

        # Default ELR TO to ELR FROM if left blank
        if not elr_to and elr_from:
            elr_to = elr_from

        search = st.button("🔍  SEARCH WORKSITE")

        # Persist search state so A&E text input doesn't reset results
        if search:
            st.session_state['ws_search'] = True
            st.session_state['ws_elr_from'] = elr_from
            st.session_state['ws_elr_to'] = elr_to
            st.session_state['ws_mil_from'] = mil_from
            st.session_state['ws_mil_to'] = mil_to

        if st.session_state.get('ws_search'):
            # Use stored values so results persist when A&E search triggers rerun
            elr_from = st.session_state.get('ws_elr_from', elr_from)
            elr_to = st.session_state.get('ws_elr_to', elr_to)
            mil_from = st.session_state.get('ws_mil_from', mil_from)
            mil_to = st.session_state.get('ws_mil_to', mil_to)

            if not elr_from:
                st.warning("Please enter an ELR FROM.")
            elif not mil_from or not mil_to:
                st.warning("Please enter mileage FROM and TO.")
            else:
                from_dec = mileage_to_decimal(mil_from)
                to_dec = mileage_to_decimal(mil_to)

                if from_dec is None or to_dec is None:
                    st.error("Invalid mileage format. Use e.g. 182m 10ch")
                else:
                    elr_label = elr_from if elr_from == elr_to else f"{elr_from} to {elr_to}"
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M")

                    # ─── SECTION 1: HAZARDS ──────────────────────────────────
                    st.markdown(f'<div class="section-header">⚠️  HAZARDS — {elr_label} {mil_from} to {mil_to}</div>',
                                unsafe_allow_html=True)

                    all_in_range = query_by_elr_mileage(hazard_df, elr_from, elr_to, from_dec, to_dec)
                    hazards = filter_hazards_only(all_in_range) if not all_in_range.empty else pd.DataFrame()

                    if hazards.empty:
                        st.markdown(f"""
                        <div class="pps-card pps-card-green">
                          <span class="badge badge-green">&#10003; Clear</span>
                          <span style="margin-left:1rem;font-size:0.85rem">No hazards found</span>
                        </div>
                        """, unsafe_allow_html=True)
                    else:
                        display_cols = ['ELR', 'ELR Name', 'Mileage  From',
                                        'Mileage To', 'Hazard Description',
                                        'Local Name', 'Track', 'Free Text']
                        hz_display = hazards[
                            [c for c in display_cols if c in hazards.columns]
                        ].fillna('')

                        mc1, mc2 = st.columns([1, 5])
                        with mc1:
                            st.markdown(f"""
                            <div class="metric-box">
                              <div class="metric-num" style="color:{COLOURS['amber']}">{len(hz_display)}</div>
                              <div class="metric-lbl">Hazards</div>
                            </div>""", unsafe_allow_html=True)

                        st.dataframe(hz_display, use_container_width=True, hide_index=True)

                        pdf_buf = generate_pdf(hz_display, elr_label, mil_from, mil_to,
                                               HAZARD_COLS, 'NWR Hazard Directory')
                        st.download_button(
                            "⬇  Download Hazard PDF", data=pdf_buf,
                            file_name=f"Hazards_{elr_from}_{timestamp}.pdf",
                            mime="application/pdf", key="hz_pdf")

                    # ─── SECTION 2: ACCESS POINTS ────────────────────────────
                    st.markdown(f'<div class="section-header">🚪  ACCESS POINTS — {elr_label} {mil_from} to {mil_to}</div>',
                                unsafe_allow_html=True)

                    ap_all = filter_access_points(hazard_df)
                    access_pts = query_by_elr_mileage(ap_all, elr_from, elr_to, from_dec, to_dec) if not ap_all.empty else pd.DataFrame()

                    if access_pts.empty:
                        st.markdown(f"""
                        <div class="pps-card pps-card-amber">
                          <span class="badge badge-amber">No access points</span>
                          <span style="margin-left:1rem;font-size:0.85rem">No access points found in range</span>
                        </div>
                        """, unsafe_allow_html=True)
                    else:
                        display_cols_ap = ['ELR', 'ELR Name', 'Mileage  From',
                                           'Hazard Description', 'Local Name',
                                           'Track', 'Free Text']
                        ap_display = access_pts[
                            [c for c in display_cols_ap if c in access_pts.columns]
                        ].fillna('')

                        # Enrich with coordinates and Google Maps links
                        ap_display = enrich_access_points_with_coords(ap_display, ap_coords_df)

                        ap_show = ap_display.rename(columns={
                            'Mileage  From': 'Mileage',
                            'Hazard Description': 'Type',
                            'Free Text': 'Details',
                        })
                        ap_show['Mileage'] = ap_show['Mileage'].apply(decimal_to_miles_chains)

                        # Also convert mileage for PDF
                        ap_display['Mileage  From'] = ap_display['Mileage  From'].apply(decimal_to_miles_chains)

                        n_ped = len(access_pts[access_pts['Hazard Description'].str.contains('Pedestrian', na=False)])
                        n_veh = len(access_pts[access_pts['Hazard Description'].str.contains('Vehicle', na=False)])
                        n_rr = len(access_pts[access_pts['Hazard Description'].str.contains('Road-Rail|Machines', na=False)])
                        n_coords = len(ap_display[ap_display['Google Maps'] != ''])

                        mc1, mc2, mc3, mc4, mc5 = st.columns(5)
                        with mc1:
                            st.markdown(f"""
                            <div class="metric-box">
                              <div class="metric-num" style="color:{COLOURS['green']}">{len(access_pts)}</div>
                              <div class="metric-lbl">Total</div>
                            </div>""", unsafe_allow_html=True)
                        with mc2:
                            st.markdown(f"""
                            <div class="metric-box">
                              <div class="metric-num" style="color:{COLOURS['text']}">{n_ped}</div>
                              <div class="metric-lbl">Pedestrian</div>
                            </div>""", unsafe_allow_html=True)
                        with mc3:
                            st.markdown(f"""
                            <div class="metric-box">
                              <div class="metric-num" style="color:{COLOURS['amber']}">{n_veh}</div>
                              <div class="metric-lbl">Vehicle</div>
                            </div>""", unsafe_allow_html=True)
                        with mc4:
                            st.markdown(f"""
                            <div class="metric-box">
                              <div class="metric-num" style="color:{COLOURS['red']}">{n_rr}</div>
                              <div class="metric-lbl">Road-Rail</div>
                            </div>""", unsafe_allow_html=True)
                        with mc5:
                            st.markdown(f"""
                            <div class="metric-box">
                              <div class="metric-num" style="color:{COLOURS['nwr_blue']}">{n_coords}</div>
                              <div class="metric-lbl">Map Links</div>
                            </div>""", unsafe_allow_html=True)

                        # On-screen: show clickable Google Maps links
                        ap_screen = ap_show.copy()
                        if 'Google Maps' in ap_screen.columns:
                            ap_screen['Location'] = ap_screen['Google Maps'].apply(
                                lambda x: f'[Map]({x})' if x else '')
                            ap_screen = ap_screen.drop(columns=['Google Maps', 'Lat', 'Lon'])

                        st.dataframe(ap_screen, use_container_width=True, hide_index=True,
                                     column_config={'Location': st.column_config.LinkColumn('Location', display_text='📍 Map')})

                        pdf_buf = generate_pdf(ap_display, elr_label, mil_from, mil_to,
                                               ACCESS_COLS, 'Access Points')
                        st.download_button(
                            "⬇  Download Access Points PDF", data=pdf_buf,
                            file_name=f"Access_Points_{elr_from}_{timestamp}.pdf",
                            mime="application/pdf", key="ap_pdf")

                    # ─── SECTION 3: SIGNAL BOX CONTACTS & LINE NAMES ─────────
                    st.markdown(f'<div class="section-header">📞  SIGNAL BOX CONTACTS — {elr_label} {mil_from} to {mil_to}</div>',
                                unsafe_allow_html=True)

                    if signalbox_df is None or signalbox_df.empty:
                        st.markdown(f"""
                        <div class="pps-card pps-card-grey">
                          <span class="badge badge-grey">Not available</span>
                          <span style="margin-left:1rem;font-size:0.85rem">
                            Place <b>signal_box_contacts.csv</b> in the data subfolder.
                          </span>
                        </div>
                        """, unsafe_allow_html=True)
                    else:
                        # Convert mileage to chains for signal box lookup
                        from_ch = int(from_dec) * 80 + round((from_dec - int(from_dec)) * 10000 / 22)
                        to_ch = int(to_dec) * 80 + round((to_dec - int(to_dec)) * 10000 / 22)

                        # Auto-detect signal boxes from mileage
                        auto_boxes = find_signal_boxes_for_mileage(
                            elr_from, elr_to, from_ch, to_ch, sba_df, signalbox_df)

                        if auto_boxes:
                            st.markdown(f"""
                            <div class="metric-box" style="max-width:200px;margin-bottom:0.8rem">
                              <div class="metric-num" style="color:{COLOURS['green']}">{len(auto_boxes)}</div>
                              <div class="metric-lbl">Signal boxes found</div>
                            </div>""", unsafe_allow_html=True)

                            for box in auto_boxes:
                                phone_display = f" — <b>{box['Phone']}</b>" if box['Phone'] else ""
                                prefix_display = f" ({box['Prefix']})" if box['Prefix'] else ""
                                # Build ECR display with phone if available
                                eco_parts = []
                                if box.get('ECO'):
                                    eco_label = box['ECO']
                                    eco_phone = format_phone_numbers(box.get('ECO Phone', ''))
                                    if eco_phone:
                                        eco_parts.append(f"ECR: {eco_label} — <b>{eco_phone}</b>")
                                    else:
                                        eco_parts.append(f"ECR: {eco_label}")
                                eco_display = f"<br/><span style='font-size:0.78rem'>{'<br/>'.join(eco_parts)}</span>" if eco_parts else ""
                                st.markdown(f"""
                                <div class="pps-card pps-card-green">
                                  <div style="font-size:1rem;color:{COLOURS['white']};font-weight:600">
                                    {box['Signal Box']}{prefix_display}{phone_display}</div>
                                  <div style="font-size:0.82rem;color:{COLOURS['muted']};margin-top:0.2rem">
                                    {box['ELR']} {box['Coverage']}{eco_display}</div>
                                </div>
                                """, unsafe_allow_html=True)
                        else:
                            st.markdown(f"""
                            <div style="font-size:0.85rem;color:{COLOURS['muted']};margin-bottom:0.8rem">
                              No signal box areas found for this ELR/mileage — use manual search below.
                            </div>
                            """, unsafe_allow_html=True)

                        # Manual search fallback
                        sb1, sb2 = st.columns([3, 3])
                        with sb1:
                            sb_name = st.text_input("Search by name", placeholder="e.g. Rugby, Colwich, Crewe", key="sb_name")
                        with sb2:
                            sb_prefix = st.text_input("Search by signal prefix", placeholder="e.g. CE, SW, AJ", key="sb_prefix").upper()

                        if sb_name or sb_prefix:
                            results = signalbox_df.copy()

                            if sb_name:
                                results = results[results['Signal Box'].str.contains(sb_name, case=False, na=False)]
                            if sb_prefix:
                                results = results[results['Signal Prefix'].str.upper().eq(sb_prefix)]

                            if results.empty:
                                st.markdown(f"""
                                <div class="pps-card pps-card-amber">
                                  <span class="badge badge-amber">No results</span>
                                  <span style="margin-left:1rem;font-size:0.85rem">
                                    No signal boxes found matching your search.
                                  </span>
                                </div>
                                """, unsafe_allow_html=True)
                            else:
                                display_sb = results[['Signal Box', 'External Telephone', 'Signal Prefix', 'Source']].copy()
                                display_sb = display_sb.fillna('')
                                display_sb = display_sb.rename(columns={
                                    'Signal Box': 'Name',
                                    'External Telephone': 'Phone',
                                    'Signal Prefix': 'Prefix',
                                    'Source': 'Route',
                                })

                                st.dataframe(display_sb, use_container_width=True, hide_index=True)

                                search_desc = sb_name or sb_prefix
                                pdf_buf = generate_pdf(results, search_desc, '', '',
                                                       SIGNALBOX_COLS, 'Signal Box Contacts')
                                st.download_button(
                                    "⬇  Download Contacts PDF", data=pdf_buf,
                                    file_name=f"Signal_Box_Contacts_{search_desc}_{timestamp}.pdf",
                                    mime="application/pdf", key="sb_pdf")

                    # ─── SECTION 3b: LINE NAMES ──────────────────────────────
                    auto_lines = find_line_names_for_mileage(
                        elr_from, elr_to,
                        int(from_dec) * 80 + round((from_dec - int(from_dec)) * 10000 / 22),
                        int(to_dec) * 80 + round((to_dec - int(to_dec)) * 10000 / 22),
                        ln_df)

                    if auto_lines:
                        st.markdown(f'<div class="section-header">🚂  LINES AT SITE — {elr_label} {mil_from} to {mil_to}</div>',
                                    unsafe_allow_html=True)

                        st.markdown(f"""
                        <div class="metric-box" style="max-width:200px;margin-bottom:0.8rem">
                          <div class="metric-num" style="color:{COLOURS['text']}">{len(auto_lines)}</div>
                          <div class="metric-lbl">Lines</div>
                        </div>""", unsafe_allow_html=True)

                        lines_df = pd.DataFrame(auto_lines)
                        st.dataframe(lines_df, use_container_width=True, hide_index=True)
                    else:
                        st.markdown(f'<div class="section-header">🚂  LINES AT SITE — {elr_label} {mil_from} to {mil_to}</div>',
                                    unsafe_allow_html=True)
                        st.markdown(f"""
                        <div style="font-size:0.85rem;color:{COLOURS['muted']};margin-bottom:0.8rem">
                          No line name data found for this ELR/mileage.
                        </div>
                        """, unsafe_allow_html=True)

                    # ─── SECTION 4: NEAREST A&E ──────────────────────────────
                    st.markdown(f'<div class="section-header">🏥  NEAREST A&amp;E — 24hr Type 1 Emergency Departments</div>',
                                unsafe_allow_html=True)

                    if ae_df is None or ae_df.empty:
                        st.markdown(f"""
                        <div class="pps-card pps-card-grey">
                          <span class="badge badge-grey">Not available</span>
                          <span style="margin-left:1rem;font-size:0.85rem">
                            Place <b>ae_departments_type1.csv</b> in the data subfolder.
                          </span>
                        </div>
                        """, unsafe_allow_html=True)
                    else:
                        # Auto-detect location from access points coordinates file
                        auto_lat, auto_lon, auto_source = find_worksite_coords(
                            elr_from, elr_to, from_dec, to_dec, ap_coords_df)

                        if auto_lat and auto_lon:
                            nearest = find_nearest_ae(auto_lat, auto_lon, ae_df, n=3)
                            st.markdown(f"""
                            <div style="font-size:0.85rem;color:{COLOURS['muted']};margin-bottom:0.8rem">
                              Location from nearest access point: <b>{auto_source}</b>
                            </div>
                            """, unsafe_allow_html=True)

                            for i, hosp in enumerate(nearest):
                                border_col = COLOURS['green'] if i == 0 else COLOURS['border']
                                label = 'NEAREST' if i == 0 else f'#{i+1}'
                                badge_class = 'badge-green' if i == 0 else 'badge-grey'
                                st.markdown(f"""
                                <div class="pps-card" style="border-left:3px solid {border_col}">
                                  <span class="badge {badge_class}">{label} — {hosp['Distance (miles)']} miles</span>
                                  <div style="margin-top:0.5rem;font-size:1rem;color:{COLOURS['white']};font-weight:600">
                                    {hosp['Hospital']}</div>
                                  <div style="font-size:0.85rem;color:{COLOURS['muted']};margin-top:0.2rem">
                                    {hosp['Address']}, {hosp['Postcode']}</div>
                                  <div style="font-size:0.78rem;color:{COLOURS['muted']};margin-top:0.1rem">
                                    {hosp['Trust']}</div>
                                </div>
                                """, unsafe_allow_html=True)
                        else:
                            st.markdown(f"""
                            <div style="font-size:0.85rem;color:{COLOURS['muted']};margin-bottom:0.8rem">
                              No coordinate data for this ELR — search by town or hospital name below.
                            </div>
                            """, unsafe_allow_html=True)

                        # Manual search fallback / override
                        ae_search = st.text_input("Search A&E by town or hospital name",
                                                  placeholder="e.g. Derby, Reading, Warrington",
                                                  key="ae_search")

                        if ae_search:
                            matches = ae_df[
                                ae_df['hospital_name'].str.contains(ae_search, case=False, na=False) |
                                ae_df['address'].str.contains(ae_search, case=False, na=False) |
                                ae_df['trust_name'].str.contains(ae_search, case=False, na=False) |
                                ae_df['postcode'].str.contains(ae_search.upper().strip(), na=False)
                            ].copy()

                            if matches.empty:
                                st.markdown(f"""
                                <div class="pps-card pps-card-amber">
                                  <span class="badge badge-amber">No results</span>
                                  <span style="margin-left:1rem;font-size:0.85rem">
                                    Try a different town name or hospital name.
                                  </span>
                                </div>
                                """, unsafe_allow_html=True)
                            else:
                                display_ae = matches[['hospital_name', 'address', 'postcode', 'trust_name']].copy()
                                display_ae.columns = ['Hospital', 'Address', 'Postcode', 'Trust']
                                st.dataframe(display_ae, use_container_width=True, hide_index=True)

    with tab_swp:
        st.markdown(f"""
        <div class="section-header">📋  SAFE WORK PACK BUILDER</div>
        """, unsafe_allow_html=True)

        st.markdown(f"""
        <div style="font-size:0.85rem;color:{COLOURS['muted']};margin-bottom:1rem">
          Paste a row from the Possession Tracker (AI version) below. The app will auto-populate the SWP
          with lines at site, signal box contacts, ECR details, nearest A&amp;E, hazards, and access points.
        </div>
        """, unsafe_allow_html=True)

        tracker_paste = st.text_area(
            "Paste tracker row here",
            placeholder="Copy a full row from the Possession Tracker and paste it here...",
            height=100,
            key="swp_paste"
        )

        # ── Extra ELRs ──
        if 'extra_elrs' not in st.session_state:
            st.session_state['extra_elrs'] = []

        if st.button("➕  Add another ELR", key="add_extra_elr_btn"):
            st.session_state['extra_elrs'].append({'elr': '', 'from': '', 'to': ''})

        for idx, extra in enumerate(st.session_state['extra_elrs']):
            ecol1, ecol2, ecol3, ecol4 = st.columns([2, 3, 3, 1])
            with ecol1:
                st.session_state['extra_elrs'][idx]['elr'] = st.text_input(
                    "ELR", value=extra['elr'], key=f"extra_elr_{idx}", label_visibility="collapsed",
                    placeholder="ELR")
            with ecol2:
                st.session_state['extra_elrs'][idx]['from'] = st.text_input(
                    "From", value=extra['from'], key=f"extra_from_{idx}", label_visibility="collapsed",
                    placeholder="From e.g. 182m 10ch")
            with ecol3:
                st.session_state['extra_elrs'][idx]['to'] = st.text_input(
                    "To", value=extra['to'], key=f"extra_to_{idx}", label_visibility="collapsed",
                    placeholder="To e.g. 183m 05ch")
            with ecol4:
                if st.button("✕", key=f"remove_extra_elr_{idx}"):
                    st.session_state['extra_elrs'].pop(idx)
                    st.rerun()

        # ── Daily List upload ──
        daily_list_file = st.file_uploader("📋  Add Daily List (NR3180)", type=["pdf"], key="daily_list_upload")
        if daily_list_file is not None:
            import fitz as _fitz
            dl_bytes = daily_list_file.read()
            st.session_state['daily_list_pdf'] = dl_bytes
            try:
                dl_doc = _fitz.open(stream=dl_bytes, filetype="pdf")
                dl_text = dl_doc[0].get_text() if len(dl_doc) > 0 else ""
                dl_doc.close()
                # Extract signal references: 2-3 uppercase letters followed by digits
                dl_signals = sorted(set(re.findall(r'\b([A-Z]{2,3}\d{2,5})\b', dl_text)))
                # Filter out common false positives
                dl_signals = [s for s in dl_signals if not s.startswith('NR') and not s.startswith('NW') and not s.startswith('P20')]
                st.session_state['daily_list_signals'] = dl_signals
                if dl_signals:
                    st.success(f"✓ Daily list loaded — protecting signals: {', '.join(dl_signals)}")
                else:
                    st.warning("Daily list loaded but no signal references found on page 1.")
            except Exception as e:
                st.error(f"Error reading daily list: {e}")
                st.session_state['daily_list_signals'] = []
        else:
            if 'daily_list_signals' not in st.session_state:
                st.session_state['daily_list_signals'] = None
            if 'daily_list_pdf' not in st.session_state:
                st.session_state['daily_list_pdf'] = None

        build_swp = st.button("📋  BUILD SWP", key="build_swp_btn")

        if build_swp and tracker_paste and tracker_paste.strip():
            st.session_state['swp_data'] = tracker_paste.strip()

        if st.session_state.get('swp_data'):
            # Parse the pasted tracker row (tab-separated)
            fields = st.session_state['swp_data'].split('\t')

            # AI Version Tracker columns:
            # 0:Week, 1:Worksite, 2:W/S Status, 3:Poss'n Ref, 4:Worktype,
            # 5:From Date, 6:From Time, 7:To Date, 8:To Time,
            # 9:ELR1, 10:Distance From 1, 11:Distance To 1,
            # 12:ELR2, 13:Distance From 2, 14:Distance To 2,
            # 15:LOR, 16:Isol, 17:ES Primary, 18:Location, 19:Client,
            # 20:Responsible Manager, 21:Planner,
            # 22:Isolation Y/N, 23:SSOW In Use, 24:Protection,
            # 25:SSOW Returned, 26:COSS Name, 27:Email, 28:Number,
            # 29:Item No, 30:Picop Meeting, 31:Planner Comments,
            # 32:SWP Ref, 33:PO Number

            def safe_get(lst, idx, default=''):
                try:
                    val = lst[idx].strip() if idx < len(lst) else default
                    # Strip surrounding quotes that come from Excel copy/paste
                    if val:
                        val = val.strip('"').strip("'").strip()
                    return val if val else default
                except (IndexError, AttributeError):
                    return default

            swp_week = safe_get(fields, 0)
            swp_worksite = safe_get(fields, 1)
            swp_status = safe_get(fields, 2)
            swp_poss_ref = safe_get(fields, 3)
            swp_worktype = safe_get(fields, 4)
            swp_from_date = safe_get(fields, 5)
            swp_from_time = safe_get(fields, 6)
            swp_to_date = safe_get(fields, 7)
            swp_to_time = safe_get(fields, 8)
            swp_elr1 = safe_get(fields, 9).upper()
            swp_dist_from1 = safe_get(fields, 10)
            swp_dist_to1 = safe_get(fields, 11)
            swp_elr2 = safe_get(fields, 12).upper()
            swp_dist_from2 = safe_get(fields, 13)
            swp_dist_to2 = safe_get(fields, 14)
            swp_lor = safe_get(fields, 15)
            swp_isol = safe_get(fields, 16)
            swp_es_primary = safe_get(fields, 17)
            swp_location = safe_get(fields, 18)
            swp_client = safe_get(fields, 19)
            swp_rm = safe_get(fields, 20)
            swp_planner = safe_get(fields, 21)
            swp_isol_yn = safe_get(fields, 22)
            swp_ssow = safe_get(fields, 23)
            swp_protection = safe_get(fields, 24)
            swp_ssow_returned = safe_get(fields, 25)
            swp_coss_name = safe_get(fields, 26)
            swp_coss_email = safe_get(fields, 27)
            swp_coss_number = safe_get(fields, 28)
            swp_item_no = safe_get(fields, 29)
            swp_picop = safe_get(fields, 30)
            swp_comments = safe_get(fields, 31)
            swp_ref = safe_get(fields, 32)
            swp_po = safe_get(fields, 33)

            # If ELR2 is blank, use ELR1 only
            swp_elr_from = swp_elr1
            swp_elr_to = swp_elr2 if swp_elr2 else swp_elr1

            # Parse RM name and number from combined field "Steve Carroll - 07939 393913"
            def parse_name_number(combined):
                if not combined:
                    return '', ''
                parts = re.split(r'\s*[-–]\s*(?=\d)', combined, maxsplit=1)
                if len(parts) == 2:
                    return parts[0].strip(), parts[1].strip()
                return combined.strip(), ''

            swp_rm_name, swp_rm_number = parse_name_number(swp_rm)
            swp_planner_name, swp_planner_number = parse_name_number(swp_planner)

            # Convert Excel date serial numbers
            def excel_date(serial):
                try:
                    serial = float(serial)
                    if serial > 40000:
                        from datetime import datetime, timedelta
                        base = datetime(1899, 12, 30)
                        return (base + timedelta(days=serial)).strftime('%d/%m/%Y')
                except (ValueError, TypeError):
                    pass
                return serial

            def excel_time(serial):
                try:
                    serial = float(serial)
                    if 0 <= serial <= 1:
                        hours = int(serial * 24)
                        mins = int((serial * 24 - hours) * 60)
                        return f"{hours:02d}:{mins:02d}"
                except (ValueError, TypeError):
                    pass
                return serial

            swp_from_date = excel_date(swp_from_date)
            swp_to_date = excel_date(swp_to_date)
            swp_from_time = excel_time(swp_from_time)
            swp_to_time = excel_time(swp_to_time)

            # Build location string
            elr_location = f"ELR: {swp_elr1} {swp_dist_from1}-{swp_dist_to1}"
            if swp_elr2:
                elr_location += f" / {swp_elr2} {swp_dist_from2}-{swp_dist_to2}"

            # ══════════════════════════════════════════════════════
            # STEP 1: PARSED JOB DETAILS (auto from tracker)
            # ══════════════════════════════════════════════════════
            st.markdown(f'<div class="section-header">📝  STEP 1 — JOB DETAILS (auto-populated)</div>', unsafe_allow_html=True)

            col1, col2 = st.columns(2)
            with col1:
                st.markdown(f"""
                <div class="pps-card pps-card-green">
                  <div style="font-size:1rem;color:{COLOURS['white']};font-weight:600;margin-bottom:0.5rem">
                    {swp_ref if swp_ref else 'No SWP Ref'}</div>
                  <div style="font-size:0.82rem;color:{COLOURS['muted']}">
                    <b>Week:</b> {swp_week}<br/>
                    <b>Worksite:</b> {swp_worksite}<br/>
                    <b>Poss Ref:</b> {swp_poss_ref}<br/>
                    <b>Work Type:</b> {swp_worktype}<br/>
                    <b>Location:</b> {swp_location}<br/>
                    <b>{elr_location}</b><br/>
                    <b>LOR:</b> {swp_lor}<br/>
                    <b>Dates:</b> {swp_from_date} {swp_from_time} — {swp_to_date} {swp_to_time}
                  </div>
                </div>
                """, unsafe_allow_html=True)
            with col2:
                st.markdown(f"""
                <div class="pps-card pps-card-green">
                  <div style="font-size:1rem;color:{COLOURS['white']};font-weight:600;margin-bottom:0.5rem">
                    Contacts & Protection</div>
                  <div style="font-size:0.82rem;color:{COLOURS['muted']}">
                    <b>Client:</b> {swp_client}<br/>
                    <b>RM:</b> {swp_rm_name} — {swp_rm_number}<br/>
                    <b>Planner:</b> {swp_planner_name} — {swp_planner_number}<br/>
                    <b>COSS:</b> {swp_coss_name}<br/>
                    <b>COSS Phone:</b> {swp_coss_number}<br/>
                    <b>COSS Email:</b> {swp_coss_email}<br/>
                    <b>Protection:</b> {swp_protection}<br/>
                    <b>SSOW:</b> {swp_ssow}<br/>
                    <b>Isolation:</b> {swp_isol} ({swp_isol_yn})
                  </div>
                </div>
                """, unsafe_allow_html=True)

            # ══════════════════════════════════════════════════════
            # AUTO-LOOKUP: Get worksite intelligence data
            # ══════════════════════════════════════════════════════
            swp_from_dec = mileage_to_decimal(swp_dist_from1)
            swp_to_dec = mileage_to_decimal(swp_dist_to1)

            swp_lines = []
            swp_boxes = []
            swp_nearest_ae = []
            swp_hazards_df = pd.DataFrame()
            swp_access_df = pd.DataFrame()

            if swp_from_dec is not None and swp_to_dec is not None:
                swp_from_ch = int(swp_from_dec) * 80 + round((swp_from_dec - int(swp_from_dec)) * 10000 / 22)
                swp_to_ch = int(swp_to_dec) * 80 + round((swp_to_dec - int(swp_to_dec)) * 10000 / 22)

                swp_lines = find_line_names_for_mileage(swp_elr_from, swp_elr_to, swp_from_ch, swp_to_ch, ln_df)
                swp_boxes = find_signal_boxes_for_mileage(swp_elr_from, swp_elr_to, swp_from_ch, swp_to_ch, sba_df, signalbox_df)

                if ae_df is not None and not ae_df.empty:
                    swp_lat, swp_lon, swp_source = find_worksite_coords(
                        swp_elr_from, swp_elr_to, swp_from_dec, swp_to_dec, ap_coords_df)
                    if swp_lat and swp_lon:
                        swp_nearest_ae = find_nearest_ae(swp_lat, swp_lon, ae_df, n=3)

                swp_hazards_df = filter_hazards_only(query_by_elr_mileage(hazard_df, swp_elr_from, swp_elr_to, swp_from_dec, swp_to_dec))
                swp_access_df = filter_access_points(query_by_elr_mileage(hazard_df, swp_elr_from, swp_elr_to, swp_from_dec, swp_to_dec))
                if not swp_access_df.empty:
                    swp_access_df = enrich_access_points_with_coords(swp_access_df, ap_coords_df)

            # ── Extra ELR lookups ──
            for extra in st.session_state.get('extra_elrs', []):
                ex_elr = extra.get('elr', '').strip().upper()
                ex_from_str = extra.get('from', '').strip()
                ex_to_str = extra.get('to', '').strip()
                if not ex_elr or not ex_from_str or not ex_to_str:
                    continue
                ex_from_dec = mileage_to_decimal(ex_from_str)
                ex_to_dec = mileage_to_decimal(ex_to_str)
                if ex_from_dec is None or ex_to_dec is None:
                    continue
                ex_from_ch = int(ex_from_dec) * 80 + round((ex_from_dec - int(ex_from_dec)) * 10000 / 22)
                ex_to_ch = int(ex_to_dec) * 80 + round((ex_to_dec - int(ex_to_dec)) * 10000 / 22)
                # Line names
                ex_lines = find_line_names_for_mileage(ex_elr, ex_elr, ex_from_ch, ex_to_ch, ln_df)
                seen_keys = {f"{l['Abbreviation']}_{l['Line Name']}" for l in swp_lines}
                for l in ex_lines:
                    key = f"{l['Abbreviation']}_{l['Line Name']}"
                    if key not in seen_keys:
                        swp_lines.append(l)
                        seen_keys.add(key)
                # Signal boxes
                ex_boxes = find_signal_boxes_for_mileage(ex_elr, ex_elr, ex_from_ch, ex_to_ch, sba_df, signalbox_df)
                seen_sb = {b['Signal Box'] for b in swp_boxes}
                for b in ex_boxes:
                    if b['Signal Box'] not in seen_sb:
                        swp_boxes.append(b)
                        seen_sb.add(b['Signal Box'])
                # Hazards and access points
                ex_hz = filter_hazards_only(query_by_elr_mileage(hazard_df, ex_elr, ex_elr, ex_from_dec, ex_to_dec))
                if not ex_hz.empty:
                    swp_hazards_df = pd.concat([swp_hazards_df, ex_hz], ignore_index=True).drop_duplicates()
                ex_ap = filter_access_points(query_by_elr_mileage(hazard_df, ex_elr, ex_elr, ex_from_dec, ex_to_dec))
                if not ex_ap.empty:
                    ex_ap = enrich_access_points_with_coords(ex_ap, ap_coords_df)
                    swp_access_df = pd.concat([swp_access_df, ex_ap], ignore_index=True).drop_duplicates()

            # ── Shift Contacts (editable) ──
            st.markdown(f'<div class="section-header" style="font-size:0.95rem">👷  Shift Contact Numbers</div>', unsafe_allow_html=True)
            st.markdown(f"<div style='font-size:0.82rem;color:{COLOURS['muted']};margin-bottom:0.5rem'>COSS auto-populated. Add PICOP, ES, and other personnel below.</div>", unsafe_allow_html=True)

            shift_contacts = []
            # Row 1: COSS (pre-filled)
            shc1, shc2, shc3, shc4, shc5 = st.columns([3, 2, 2, 1, 1])
            with shc1:
                sc_name_0 = st.text_input("Name", value=swp_coss_name, key="swp_sc_name_0")
            with shc2:
                sc_duty_0 = st.text_input("Duty", value="COSS", key="swp_sc_duty_0")
            with shc3:
                sc_phone_0 = st.text_input("Phone", value=swp_coss_number, key="swp_sc_phone_0")
            with shc4:
                sc_start_0 = st.text_input("Start", value=swp_from_time, key="swp_sc_start_0")
            with shc5:
                sc_end_0 = st.text_input("End", value=swp_to_time, key="swp_sc_end_0")
            shift_contacts.append({'Name': sc_name_0, 'Duty': sc_duty_0, 'Phone': sc_phone_0, 'Start': sc_start_0, 'End': sc_end_0})

            # Rows 2-5: blank for PICOP, ES, etc
            for i in range(1, 5):
                shc1, shc2, shc3, shc4, shc5 = st.columns([3, 2, 2, 1, 1])
                with shc1:
                    sc_name = st.text_input("Name", value="", key=f"swp_sc_name_{i}", label_visibility="collapsed")
                with shc2:
                    sc_duty = st.text_input("Duty", value="", key=f"swp_sc_duty_{i}", label_visibility="collapsed")
                with shc3:
                    sc_phone = st.text_input("Phone", value="", key=f"swp_sc_phone_{i}", label_visibility="collapsed")
                with shc4:
                    sc_start = st.text_input("Start", value="", key=f"swp_sc_start_{i}", label_visibility="collapsed")
                with shc5:
                    sc_end = st.text_input("End", value="", key=f"swp_sc_end_{i}", label_visibility="collapsed")
                if sc_name.strip():
                    shift_contacts.append({'Name': sc_name, 'Duty': sc_duty, 'Phone': sc_phone, 'Start': sc_start, 'End': sc_end})

            # ══════════════════════════════════════════════════════
            # STEP 2: EDITABLE FIELDS
            # ══════════════════════════════════════════════════════
            st.markdown(f'<div class="section-header">✏️  STEP 2 — COMPLETE THE PACK</div>', unsafe_allow_html=True)

            # ── Nature of Work (editable, pre-filled, overrides Page 1) ──
            swp_nature_of_work = st.text_area(
                "Nature of Work (expand if needed — this overrides Page 1)",
                value=swp_worktype,
                height=80,
                key="swp_nature"
            )

            # ── SWP Type: Cyclical / Non-Cyclical / Repeat ──
            swp_type = st.selectbox(
                "SWP Type",
                ["Cyclical", "Non-Cyclical", "Repeat"],
                index=1,
                key="swp_type_sel"
            )

            # ── Protection Method ──
            st.markdown(f'<div class="section-header" style="font-size:0.95rem">🛡️  Protection Method</div>', unsafe_allow_html=True)

            protection_options = [
                "1 - Safeguarded site of work",
                "2 - Fenced site of work",
                "3 - Separated site of work",
                "4 - Warning systems - permanent train activated",
                "5 - Warning systems - portable train activated",
                "6 - Lookout warning (max 25mph)"
            ]

            # Pre-select from tracker
            default_prot_idx = 0
            prot_lower = swp_protection.lower() if swp_protection else ''
            if 'safeguard' in prot_lower:
                default_prot_idx = 0
            elif 'fenced' in prot_lower or 'fence' in prot_lower:
                default_prot_idx = 1
            elif 'separat' in prot_lower:
                default_prot_idx = 2
            elif 'possession' in prot_lower:
                default_prot_idx = 0
            elif 'lineblock' in prot_lower or 'line block' in prot_lower:
                default_prot_idx = 2
            elif 'site warden' in prot_lower:
                default_prot_idx = 2

            swp_prot_selected = st.selectbox(
                "Protection method selected",
                protection_options,
                index=default_prot_idx,
                key="swp_protection_sel"
            )

            # If not safeguarded, show reason fields
            prot_num = int(swp_prot_selected[0])

            sg_reason = ''
            fence_reason = ''
            if prot_num >= 2:
                sg_reason = st.text_input(
                    "Reason Safeguarded NOT selected",
                    value="Not Available",
                    key="swp_sg_reason"
                )
            if prot_num >= 3:
                fence_reason = st.text_input(
                    "Reason Fenced NOT selected",
                    value="Disproportionate 25% rule",
                    key="swp_fence_reason"
                )

            # If fenced — fence details
            swp_fence_type = ""
            swp_fence_dist = ""
            if prot_num == 2:
                fc1, fc2 = st.columns(2)
                with fc1:
                    swp_fence_type = st.selectbox("Type of fence", ["Vortok", "Blue Netlon", "Black & Yellow Tape", "N/A"], key="swp_fence_type")
                with fc2:
                    swp_fence_dist = st.selectbox("Distance from line", ["1.25m", "2m", "N/A"], key="swp_fence_dist")

            # If separated — separation details
            swp_sep_dist = ""
            swp_sep_warning = ""
            if prot_num == 3:
                sc1, sc2 = st.columns(2)
                with sc1:
                    swp_sep_dist = st.selectbox("Separation distance", ["2m", "N/A"], key="swp_sep_dist")
                with sc2:
                    swp_sep_warning = st.selectbox("How Site Warden will give warning", ["Verbal", "Horn", "Whistle", "Horn or Whistle", "N/A"], key="swp_sep_warn")

            # SSOW walking vs working tick boxes
            st.markdown(f'<div class="section-header" style="font-size:0.95rem">🚶  SSOW — Walking to Site vs Working</div>', unsafe_allow_html=True)

            ssow_col1, ssow_col2 = st.columns(2)
            with ssow_col1:
                swp_ssow_walking = st.selectbox(
                    "SSOW for walking to/from site",
                    protection_options,
                    index=default_prot_idx,
                    key="swp_ssow_walking"
                )
            with ssow_col2:
                swp_ssow_working = st.selectbox(
                    "SSOW for working at site",
                    protection_options,
                    index=default_prot_idx,
                    key="swp_ssow_working"
                )

            # ── Runaway Risk Analysis ──
            st.markdown(f'<div class="section-header" style="font-size:0.95rem">⚠️  Runaway Risk Analysis</div>', unsafe_allow_html=True)

            runaway_questions = [
                "Are the works taking place on or near the line?",
                "Could your work potentially lead to a Runaway?",
                "Are my works within a Possession or adjacent to a Possession?",
                "Are my works on a gradient steeper than 1 in 100 or is there a gradient within 5 miles?",
                "Is the site of work at risk of a runaway from a 3rd Party?"
            ]
            runaway_defaults = ["Yes", "No", "No", "No", "No"]

            runaway_answers = []
            runaway_comments = []
            for i, q in enumerate(runaway_questions):
                rc1, rc2, rc3 = st.columns([4, 1, 3])
                with rc1:
                    st.markdown(f"<div style='font-size:0.82rem;color:{COLOURS['text']};padding-top:0.5rem'>{q}</div>", unsafe_allow_html=True)
                with rc2:
                    ans = st.selectbox("", ["Yes", "No"], index=0 if runaway_defaults[i] == "Yes" else 1, key=f"swp_ra_{i}", label_visibility="collapsed")
                    runaway_answers.append(ans)
                with rc3:
                    comment = st.text_input("", placeholder="Comment...", key=f"swp_ra_c_{i}", label_visibility="collapsed")
                    runaway_comments.append(comment)

            # ── Lines at Site ──
            st.markdown(f'<div class="section-header" style="font-size:0.95rem">🚂  Lines at Site</div>', unsafe_allow_html=True)

            swp_line_data = []
            if swp_lines:
                for i, line in enumerate(swp_lines):
                    lc0, lc1, lc2, lc3 = st.columns([0.3, 2, 3, 2])
                    with lc0:
                        include = st.checkbox("", value=True, key=f"swp_ln_inc_{i}", label_visibility="collapsed")
                    with lc1:
                        abbr = st.text_input("Abbrev", value=line['Abbreviation'], key=f"swp_ln_a_{i}", label_visibility="collapsed")
                    with lc2:
                        name = st.text_input("Line Name", value=line['Line Name'], key=f"swp_ln_n_{i}", label_visibility="collapsed")
                    with lc3:
                        status = st.selectbox("Status", ["Open", "Blocked", "Blocked in between trains"], key=f"swp_ln_s_{i}", label_visibility="collapsed")
                    if include:
                        swp_line_data.append({'Abbreviation': abbr, 'Line Name': name, 'Status': status})
            else:
                st.markdown(f"<div style='font-size:0.85rem;color:{COLOURS['muted']}'>No line data found. Add manually below.</div>", unsafe_allow_html=True)

            # Add extra line option
            if st.checkbox("Add extra line manually", key="swp_add_line"):
                elc1, elc2, elc3 = st.columns([2, 3, 2])
                with elc1:
                    extra_abbr = st.text_input("Abbrev", key="swp_ln_extra_a", label_visibility="collapsed")
                with elc2:
                    extra_name = st.text_input("Line Name", key="swp_ln_extra_n", label_visibility="collapsed")
                with elc3:
                    extra_status = st.selectbox("Status", ["Open", "Blocked", "Blocked in between trains"], key="swp_ln_extra_s", label_visibility="collapsed")
                if extra_abbr:
                    swp_line_data.append({'Abbreviation': extra_abbr, 'Line Name': extra_name, 'Status': extra_status})

            # ── Task Key Risks ──
            st.markdown(f'<div class="section-header" style="font-size:0.95rem">📋  Task Key Risks & Controls</div>', unsafe_allow_html=True)

            swp_task_risks = st.text_area(
                "Task Key Risks and Controls",
                value="See Task Briefing (NWR Life Saving Rules Apply)",
                height=80,
                key="swp_task_risks"
            )

            # ── Permits ──
            swp_permits = st.text_input(
                "Permits Required",
                value="None required",
                key="swp_permits"
            )

            # ── Access & Egress ──
            st.markdown(f'<div class="section-header" style="font-size:0.95rem">🚪  Access & Egress</div>', unsafe_allow_html=True)

            access_options = []
            if not swp_access_df.empty:
                for _, row in swp_access_df.iterrows():
                    local = str(row.get('Local Name', '')).strip()
                    mil = str(row.get('Mileage  From', '')).strip()
                    desc = str(row.get('Hazard Description', '')).strip()
                    elr = str(row.get('ELR', '')).strip()
                    free_text = str(row.get('Free Text', '')).strip()
                    if local and local != 'nan':
                        label = f"{local} ({elr} {mil})"
                    elif desc and desc != 'nan':
                        label = f"{desc} ({elr} {mil})"
                    else:
                        label = f"{elr} {mil}"
                    access_options.append(label)

            if access_options:
                swp_access_selected = st.multiselect(
                    "Select access point(s) from database",
                    access_options,
                    default=[access_options[0]] if access_options else [],
                    key="swp_access_sel"
                )
            else:
                swp_access_selected = []
                st.markdown(f"<div style='font-size:0.85rem;color:{COLOURS['muted']}'>No access points found in database.</div>", unsafe_allow_html=True)

            swp_access_manual = st.text_input(
                "Additional access details (if not in database)",
                placeholder="e.g. Access via field gate at 131m 02ch, postcode DE24 8EB",
                key="swp_access_manual"
            )

            # ── Hazards (both fields) ──
            st.markdown(f'<div class="section-header" style="font-size:0.95rem">⚠️  Hazards</div>', unsafe_allow_html=True)

            swp_access_hazards = st.text_area(
                "Hazards associated with access/egress",
                value="*** See Hazard Directory ***",
                height=60,
                key="swp_access_hazards"
            )

            swp_site_hazards = st.text_area(
                "Hazards associated with the site",
                value="*** See Hazard Directory ***",
                height=60,
                key="swp_site_hazards"
            )

            # ── Limits of Working Area ──
            swp_limits = st.text_input(
                "Limits of the working area",
                value=f"{elr_location}",
                key="swp_limits"
            )

            # ── Planner Comments / Instructions ──
            st.markdown(f'<div class="section-header" style="font-size:0.95rem">📝  Planner Comments / Instructions</div>', unsafe_allow_html=True)
            swp_planner_comments = st.text_area(
                "Additional comments or instructions for the COSS (shown in yellow on Page 1)",
                value=swp_comments if swp_comments else "",
                height=80,
                key="swp_planner_comments"
            )

            # ══════════════════════════════════════════════════════
            # STEP 3: AUTO-POPULATED DATA (read-only display)
            # ══════════════════════════════════════════════════════
            st.markdown(f'<div class="section-header">📊  STEP 3 — AUTO-POPULATED DATA</div>', unsafe_allow_html=True)

            # ── Signal Box Contacts ──
            st.markdown(f'<div class="section-header" style="font-size:0.95rem">📞  Signal Box Contacts</div>', unsafe_allow_html=True)
            if swp_boxes:
                for box in swp_boxes:
                    phone_display = f" — <b>{box['Phone']}</b>" if box['Phone'] else ""
                    prefix_display = f" ({box['Prefix']})" if box['Prefix'] else ""
                    eco_parts = []
                    if box.get('ECO'):
                        eco_label = box['ECO']
                        eco_phone = format_phone_numbers(box.get('ECO Phone', ''))
                        if eco_phone:
                            eco_parts.append(f"ECR: {eco_label} — <b>{eco_phone}</b>")
                        else:
                            eco_parts.append(f"ECR: {eco_label}")
                    eco_display = f"<br/><span style='font-size:0.78rem'>{'<br/>'.join(eco_parts)}</span>" if eco_parts else ""
                    st.markdown(f"""
                    <div class="pps-card pps-card-green">
                      <div style="font-size:1rem;color:{COLOURS['white']};font-weight:600">
                        {box['Signal Box']}{prefix_display}{phone_display}</div>
                      <div style="font-size:0.82rem;color:{COLOURS['muted']};margin-top:0.2rem">
                        {box['ELR']} {box['Coverage']}{eco_display}</div>
                    </div>
                    """, unsafe_allow_html=True)
            else:
                st.markdown(f"<div style='font-size:0.85rem;color:{COLOURS['muted']}'>No signal box data found.</div>", unsafe_allow_html=True)

            # ── Nearest A&E ──
            st.markdown(f'<div class="section-header" style="font-size:0.95rem">🏥  Nearest A&amp;E</div>', unsafe_allow_html=True)
            if swp_nearest_ae:
                hosp = swp_nearest_ae[0]
                st.markdown(f"""
                <div class="pps-card pps-card-green">
                  <span class="badge badge-green">NEAREST — {hosp['Distance (miles)']} miles</span>
                  <div style="margin-top:0.5rem;font-size:1rem;color:{COLOURS['white']};font-weight:600">
                    {hosp['Hospital']}</div>
                  <div style="font-size:0.85rem;color:{COLOURS['muted']};margin-top:0.2rem">
                    {hosp['Address']}, {hosp['Postcode']}</div>
                </div>
                """, unsafe_allow_html=True)
            else:
                st.markdown(f"<div style='font-size:0.85rem;color:{COLOURS['muted']}'>No A&E data available.</div>", unsafe_allow_html=True)

            # ── Hazards Summary ──
            st.markdown(f'<div class="section-header" style="font-size:0.95rem">⚠️  Hazards from Directory ({len(swp_hazards_df)})</div>', unsafe_allow_html=True)
            if not swp_hazards_df.empty:
                swp_haz_display = swp_hazards_df.copy()
                if 'Mileage  From' in swp_haz_display.columns:
                    swp_haz_display['Mileage  From'] = swp_haz_display['Mileage  From'].apply(decimal_to_miles_chains)
                if 'Mileage To' in swp_haz_display.columns:
                    swp_haz_display['Mileage To'] = swp_haz_display['Mileage To'].apply(decimal_to_miles_chains)
                display_cols = [c for c in ['ELR', 'Mileage  From', 'Mileage To', 'Hazard Description', 'Local Name', 'Free Text'] if c in swp_haz_display.columns]
                st.dataframe(swp_haz_display[display_cols].fillna(''), use_container_width=True, hide_index=True, height=200)
            else:
                st.markdown(f"<div style='font-size:0.85rem;color:{COLOURS['muted']}'>No hazards found.</div>", unsafe_allow_html=True)

            # ── Access Points Summary ──
            st.markdown(f'<div class="section-header" style="font-size:0.95rem">🚪  Access Points ({len(swp_access_df)})</div>', unsafe_allow_html=True)
            if not swp_access_df.empty:
                swp_ap_display = swp_access_df.copy()
                if 'Mileage  From' in swp_ap_display.columns:
                    swp_ap_display['Mileage  From'] = swp_ap_display['Mileage  From'].apply(decimal_to_miles_chains)
                display_cols = [c for c in ['ELR', 'Mileage  From', 'Hazard Description', 'Local Name', 'Free Text', 'Google Maps'] if c in swp_ap_display.columns]
                st.dataframe(swp_ap_display[display_cols].fillna(''), use_container_width=True, hide_index=True, height=200)
            else:
                st.markdown(f"<div style='font-size:0.85rem;color:{COLOURS['muted']}'>No access points found.</div>", unsafe_allow_html=True)


            # ══════════════════════════════════════════════════════
            # GENERATE SWP
            # ══════════════════════════════════════════════════════
            st.markdown("<br/>", unsafe_allow_html=True)

            gen_col1, gen_col2, gen_col3, gen_col4 = st.columns(4)
            with gen_col1:
                generate_btn = st.button("📥  GENERATE SWP (Excel + PDF)", key="generate_swp_btn")
            with gen_col2:
                haz_pdf_btn = st.button("📋  Hazard Directory & Access Points", key="swp_haz_pdf_btn")
                if haz_pdf_btn:
                    swp_elr_label = swp_elr_from if swp_elr_from == swp_elr_to else f"{swp_elr_from} to {swp_elr_to}"
                    swp_mil_label_from = swp_dist_from1 or ''
                    swp_mil_label_to = swp_dist_to1 or ''
                    swp_ts = datetime.now().strftime("%Y%m%d_%H%M")

                    # Hazards PDF
                    if not swp_hazards_df.empty:
                        hz_pdf_df = swp_hazards_df.copy()
                        if 'Mileage  From' in hz_pdf_df.columns:
                            hz_pdf_df['Mileage  From'] = hz_pdf_df['Mileage  From'].apply(decimal_to_miles_chains)
                        if 'Mileage To' in hz_pdf_df.columns:
                            hz_pdf_df['Mileage To'] = hz_pdf_df['Mileage To'].apply(decimal_to_miles_chains)
                        display_cols = [c for c in ['ELR', 'ELR Name', 'Mileage  From', 'Mileage To',
                                                    'Hazard Description', 'Local Name', 'Track', 'Free Text']
                                        if c in hz_pdf_df.columns]
                        hz_pdf_df = hz_pdf_df[display_cols].fillna('')
                        hz_buf = generate_pdf(hz_pdf_df, swp_elr_label,
                                              swp_mil_label_from, swp_mil_label_to,
                                              HAZARD_COLS, 'NWR Hazard Directory')
                        st.download_button(
                            "⬇  Download Hazards PDF", data=hz_buf,
                            file_name=f"Hazards_{swp_elr_from}_{swp_ts}.pdf",
                            mime="application/pdf", key="swp_hz_dl")

                    # Access Points PDF
                    if not swp_access_df.empty:
                        ap_pdf_df = swp_access_df.copy()
                        if 'Mileage  From' in ap_pdf_df.columns:
                            ap_pdf_df['Mileage  From'] = ap_pdf_df['Mileage  From'].apply(decimal_to_miles_chains)
                        display_cols = [c for c in ['ELR', 'ELR Name', 'Mileage  From',
                                                    'Hazard Description', 'Local Name', 'Track',
                                                    'Free Text', 'Google Maps']
                                        if c in ap_pdf_df.columns]
                        ap_pdf_df = ap_pdf_df[display_cols].fillna('')
                        ap_buf = generate_pdf(ap_pdf_df, swp_elr_label,
                                              swp_mil_label_from, swp_mil_label_to,
                                              ACCESS_COLS, 'Access Points')
                        st.download_button(
                            "⬇  Download Access Points PDF", data=ap_buf,
                            file_name=f"Access_Points_{swp_elr_from}_{swp_ts}.pdf",
                            mime="application/pdf", key="swp_ap_dl")

                    if swp_hazards_df.empty and swp_access_df.empty:
                        st.info("No hazards or access points found for this mileage range.")
            with gen_col3:
                sa_btn = st.button("📐  Sectional Appendix", key="swp_sa_btn")
                if sa_btn:
                    # Find matching line_names rows for worksite ELR(s) and mileage
                    sa_pages = set()
                    sa_source_docs = set()
                    if ln_df is not None and not ln_df.empty and swp_from_dec is not None and swp_to_dec is not None:
                        sa_from_ch = int(swp_from_dec) * 80 + round((swp_from_dec - int(swp_from_dec)) * 10000 / 22)
                        sa_to_ch = int(swp_to_dec) * 80 + round((swp_to_dec - int(swp_to_dec)) * 10000 / 22)

                        # Build list of ELRs to query
                        sa_elrs = [swp_elr_from]
                        if swp_elr_to and swp_elr_to != swp_elr_from:
                            sa_elrs.append(swp_elr_to)
                        for extra in st.session_state.get('extra_elrs', []):
                            ex_elr = extra.get('elr', '').strip().upper()
                            if ex_elr and ex_elr not in sa_elrs:
                                sa_elrs.append(ex_elr)

                        for elr_q in sa_elrs:
                            # Use extra ELR mileage if available, otherwise main mileage
                            q_from_ch, q_to_ch = sa_from_ch, sa_to_ch
                            for extra in st.session_state.get('extra_elrs', []):
                                if extra.get('elr', '').strip().upper() == elr_q:
                                    ex_from = mileage_to_decimal(extra.get('from', ''))
                                    ex_to = mileage_to_decimal(extra.get('to', ''))
                                    if ex_from is not None and ex_to is not None:
                                        q_from_ch = int(ex_from) * 80 + round((ex_from - int(ex_from)) * 10000 / 22)
                                        q_to_ch = int(ex_to) * 80 + round((ex_to - int(ex_to)) * 10000 / 22)
                                    break

                            matches = ln_df[
                                (ln_df['elr'] == elr_q) &
                                (ln_df['mileage_from_ch'] <= q_to_ch) &
                                (ln_df['mileage_to_ch'] >= q_from_ch)
                            ]
                            for _, row in matches.iterrows():
                                if pd.notna(row.get('source_page')):
                                    sa_pages.add(int(row['source_page']))
                                if pd.notna(row.get('source_doc')):
                                    sa_source_docs.add(str(row['source_doc']))

                    if not sa_pages:
                        st.info("No Sectional Appendix pages found for this worksite.")
                    else:
                        # Map source_doc to local PDF path
                        sa_pdf_path = None
                        sa_pdf_dir = os.path.join(os.path.dirname(__file__), 'data', 'sectional_appendix')
                        for doc_name in sa_source_docs:
                            if 'London North Western (North)' in doc_name:
                                sa_pdf_path = os.path.join(sa_pdf_dir, 'London North Western (North) Sectional Appendix March 2026.pdf')
                                break

                        if sa_pdf_path is None or not os.path.exists(sa_pdf_path):
                            st.warning("Sectional Appendix PDF not available for this region.")
                        else:
                            import fitz as _fitz
                            sa_doc = _fitz.open(sa_pdf_path)
                            sa_out = _fitz.open()
                            for pg in sorted(sa_pages):
                                if 0 < pg <= len(sa_doc):
                                    sa_out.insert_pdf(sa_doc, from_page=pg-1, to_page=pg-1)
                            sa_bytes = sa_out.tobytes()
                            sa_out.close()
                            sa_doc.close()
                            sa_ts = datetime.now().strftime("%Y%m%d_%H%M")
                            st.download_button(
                                f"⬇  Download {len(sa_pages)} SA pages",
                                data=sa_bytes,
                                file_name=f"Sectional_Appendix_{swp_elr_from}_{sa_ts}.pdf",
                                mime="application/pdf",
                                key="swp_sa_dl")
            with gen_col4:
                sd_btn = st.button("🚦  Signal Diagram", key="swp_sd_btn")
                if sd_btn:
                    sd_pages = []  # list of (diagram_doc, diagram_page)
                    dl_signals = st.session_state.get('daily_list_signals')

                    if dl_signals and sig_ref_df is not None and not sig_ref_df.empty:
                        # Daily list mode: look up signal refs
                        for sig in dl_signals:
                            sig_matches = sig_ref_df[sig_ref_df['signal_ref'] == sig]
                            for _, row in sig_matches.iterrows():
                                if pd.notna(row.get('diagram_doc')) and pd.notna(row.get('diagram_page')):
                                    sd_pages.append((str(row['diagram_doc']), int(row['diagram_page'])))
                        # Fill in pages between min and max per diagram_doc
                        if sd_pages:
                            from collections import defaultdict
                            doc_pages = defaultdict(set)
                            for doc, pg in sd_pages:
                                doc_pages[doc].add(pg)
                            sd_pages = []
                            for doc, pages in doc_pages.items():
                                for pg in range(min(pages), max(pages) + 1):
                                    sd_pages.append((doc, pg))
                    elif sd_idx_df is not None and not sd_idx_df.empty and swp_from_dec is not None and swp_to_dec is not None:
                        # ELR + mileage mode (existing logic)
                        sd_from_ch = int(swp_from_dec) * 80 + round((swp_from_dec - int(swp_from_dec)) * 10000 / 22)
                        sd_to_ch = int(swp_to_dec) * 80 + round((swp_to_dec - int(swp_to_dec)) * 10000 / 22)

                        sd_elrs = [swp_elr_from]
                        if swp_elr_to and swp_elr_to != swp_elr_from:
                            sd_elrs.append(swp_elr_to)
                        for extra in st.session_state.get('extra_elrs', []):
                            ex_elr = extra.get('elr', '').strip().upper()
                            if ex_elr and ex_elr not in sd_elrs:
                                sd_elrs.append(ex_elr)

                        for elr_q in sd_elrs:
                            q_from_ch, q_to_ch = sd_from_ch, sd_to_ch
                            for extra in st.session_state.get('extra_elrs', []):
                                if extra.get('elr', '').strip().upper() == elr_q:
                                    ex_from = mileage_to_decimal(extra.get('from', ''))
                                    ex_to = mileage_to_decimal(extra.get('to', ''))
                                    if ex_from is not None and ex_to is not None:
                                        q_from_ch = int(ex_from) * 80 + round((ex_from - int(ex_from)) * 10000 / 22)
                                        q_to_ch = int(ex_to) * 80 + round((ex_to - int(ex_to)) * 10000 / 22)
                                    break
                            matches = sd_idx_df[
                                (sd_idx_df['elr'] == elr_q) &
                                (sd_idx_df['mileage_from_ch'] <= q_to_ch) &
                                (sd_idx_df['mileage_to_ch'] >= q_from_ch)
                            ]
                            for _, row in matches.iterrows():
                                sd_pages.append((str(row['diagram_doc']), int(row['diagram_page'])))

                    sd_pages = sorted(set(sd_pages))

                    if not sd_pages:
                        st.info("No Signal Diagram pages found for this worksite.")
                    else:
                        import fitz as _fitz
                        sd_dir = os.path.join(os.path.dirname(__file__), 'data', 'signal_diagrams')
                        sd_file_map = {}
                        for dirpath, _, filenames in os.walk(sd_dir):
                            for fn in filenames:
                                sd_file_map[fn] = os.path.join(dirpath, fn)

                        sd_out = _fitz.open()
                        sd_missing = set()
                        sd_extracted = 0
                        for doc_name, pg in sd_pages:
                            if doc_name not in sd_file_map:
                                sd_missing.add(doc_name)
                                continue
                            try:
                                src = _fitz.open(sd_file_map[doc_name])
                                if 0 < pg <= len(src):
                                    sd_out.insert_pdf(src, from_page=pg-1, to_page=pg-1)
                                    sd_extracted += 1
                                src.close()
                            except Exception:
                                sd_missing.add(doc_name)

                        if sd_missing:
                            st.warning(f"Missing signal diagram PDFs: {', '.join(sorted(sd_missing))}")

                        if sd_extracted > 0:
                            sd_bytes = sd_out.tobytes()
                            sd_out.close()
                            sd_ts = datetime.now().strftime("%Y%m%d_%H%M")
                            st.download_button(
                                f"⬇  Download {sd_extracted} SD pages",
                                data=sd_bytes,
                                file_name=f"Signal_Diagrams_{swp_elr_from}_{sd_ts}.pdf",
                                mime="application/pdf",
                                key="swp_sd_dl")
                        else:
                            sd_out.close()

            # ══════════════════════════════════════════════════════
            # DOWNLOAD COMPLETE PACK
            # ══════════════════════════════════════════════════════
            st.markdown("")
            pack_btn = st.button("📦  Download Complete Pack", key="swp_pack_btn")
            if pack_btn:
                if not st.session_state.get('swp_pdf_bytes'):
                    st.warning("Generate the SWP first before downloading the complete pack.")
                else:
                    import fitz as _fitz
                    pack_pdf = _fitz.open()
                    pack_ts = datetime.now().strftime("%Y%m%d_%H%M")

                    # 1. SWP PDF
                    swp_src = _fitz.open(stream=st.session_state['swp_pdf_bytes'], filetype="pdf")
                    pack_pdf.insert_pdf(swp_src)
                    swp_src.close()

                    # 2. Daily List PDF
                    dl_pdf_bytes = st.session_state.get('daily_list_pdf')
                    if dl_pdf_bytes:
                        dl_src = _fitz.open(stream=dl_pdf_bytes, filetype="pdf")
                        pack_pdf.insert_pdf(dl_src)
                        dl_src.close()

                    # 3. Sectional Appendix pages
                    if ln_df is not None and not ln_df.empty and swp_from_dec is not None and swp_to_dec is not None:
                        pk_sa_from_ch = int(swp_from_dec) * 80 + round((swp_from_dec - int(swp_from_dec)) * 10000 / 22)
                        pk_sa_to_ch = int(swp_to_dec) * 80 + round((swp_to_dec - int(swp_to_dec)) * 10000 / 22)
                        pk_sa_elrs = [swp_elr_from]
                        if swp_elr_to and swp_elr_to != swp_elr_from:
                            pk_sa_elrs.append(swp_elr_to)
                        for extra in st.session_state.get('extra_elrs', []):
                            ex_elr = extra.get('elr', '').strip().upper()
                            if ex_elr and ex_elr not in pk_sa_elrs:
                                pk_sa_elrs.append(ex_elr)
                        pk_sa_pages = set()
                        pk_sa_docs = set()
                        for elr_q in pk_sa_elrs:
                            q_from_ch, q_to_ch = pk_sa_from_ch, pk_sa_to_ch
                            for extra in st.session_state.get('extra_elrs', []):
                                if extra.get('elr', '').strip().upper() == elr_q:
                                    ex_f = mileage_to_decimal(extra.get('from', ''))
                                    ex_t = mileage_to_decimal(extra.get('to', ''))
                                    if ex_f is not None and ex_t is not None:
                                        q_from_ch = int(ex_f) * 80 + round((ex_f - int(ex_f)) * 10000 / 22)
                                        q_to_ch = int(ex_t) * 80 + round((ex_t - int(ex_t)) * 10000 / 22)
                                    break
                            for _, row in ln_df[(ln_df['elr']==elr_q) & (ln_df['mileage_from_ch']<=q_to_ch) & (ln_df['mileage_to_ch']>=q_from_ch)].iterrows():
                                if pd.notna(row.get('source_page')):
                                    pk_sa_pages.add(int(row['source_page']))
                                if pd.notna(row.get('source_doc')):
                                    pk_sa_docs.add(str(row['source_doc']))
                        if pk_sa_pages:
                            pk_sa_path = None
                            pk_sa_dir = os.path.join(os.path.dirname(__file__), 'data', 'sectional_appendix')
                            for dn in pk_sa_docs:
                                if 'London North Western (North)' in dn:
                                    pk_sa_path = os.path.join(pk_sa_dir, 'London North Western (North) Sectional Appendix March 2026.pdf')
                                    break
                            if pk_sa_path and os.path.exists(pk_sa_path):
                                sa_src = _fitz.open(pk_sa_path)
                                for pg in sorted(pk_sa_pages):
                                    if 0 < pg <= len(sa_src):
                                        pack_pdf.insert_pdf(sa_src, from_page=pg-1, to_page=pg-1)
                                sa_src.close()

                    # 4. Signal Diagram pages (daily list signals if available, else ELR+mileage)
                    pk_sd_pages = []
                    pk_dl_signals = st.session_state.get('daily_list_signals')
                    if pk_dl_signals and sig_ref_df is not None and not sig_ref_df.empty:
                        for sig in pk_dl_signals:
                            for _, row in sig_ref_df[sig_ref_df['signal_ref']==sig].iterrows():
                                if pd.notna(row.get('diagram_doc')) and pd.notna(row.get('diagram_page')):
                                    pk_sd_pages.append((str(row['diagram_doc']), int(row['diagram_page'])))
                        if pk_sd_pages:
                            from collections import defaultdict
                            doc_pgs = defaultdict(set)
                            for doc, pg in pk_sd_pages:
                                doc_pgs[doc].add(pg)
                            pk_sd_pages = []
                            for doc, pgs in doc_pgs.items():
                                for pg in range(min(pgs), max(pgs)+1):
                                    pk_sd_pages.append((doc, pg))
                    elif sd_idx_df is not None and not sd_idx_df.empty and swp_from_dec is not None and swp_to_dec is not None:
                        pk_sd_from = int(swp_from_dec) * 80 + round((swp_from_dec - int(swp_from_dec)) * 10000 / 22)
                        pk_sd_to = int(swp_to_dec) * 80 + round((swp_to_dec - int(swp_to_dec)) * 10000 / 22)
                        pk_sd_elrs = [swp_elr_from]
                        if swp_elr_to and swp_elr_to != swp_elr_from:
                            pk_sd_elrs.append(swp_elr_to)
                        for extra in st.session_state.get('extra_elrs', []):
                            ex_elr = extra.get('elr', '').strip().upper()
                            if ex_elr and ex_elr not in pk_sd_elrs:
                                pk_sd_elrs.append(ex_elr)
                        for elr_q in pk_sd_elrs:
                            q_from_ch, q_to_ch = pk_sd_from, pk_sd_to
                            for extra in st.session_state.get('extra_elrs', []):
                                if extra.get('elr', '').strip().upper() == elr_q:
                                    ex_f = mileage_to_decimal(extra.get('from', ''))
                                    ex_t = mileage_to_decimal(extra.get('to', ''))
                                    if ex_f is not None and ex_t is not None:
                                        q_from_ch = int(ex_f) * 80 + round((ex_f - int(ex_f)) * 10000 / 22)
                                        q_to_ch = int(ex_t) * 80 + round((ex_t - int(ex_t)) * 10000 / 22)
                                    break
                            for _, row in sd_idx_df[(sd_idx_df['elr']==elr_q) & (sd_idx_df['mileage_from_ch']<=q_to_ch) & (sd_idx_df['mileage_to_ch']>=q_from_ch)].iterrows():
                                pk_sd_pages.append((str(row['diagram_doc']), int(row['diagram_page'])))
                    pk_sd_pages = sorted(set(pk_sd_pages))
                    if pk_sd_pages:
                        sd_dir = os.path.join(os.path.dirname(__file__), 'data', 'signal_diagrams')
                        sd_fm = {}
                        for dp, _, fns in os.walk(sd_dir):
                            for fn in fns:
                                sd_fm[fn] = os.path.join(dp, fn)
                        for doc_name, pg in pk_sd_pages:
                            if doc_name in sd_fm:
                                try:
                                    src = _fitz.open(sd_fm[doc_name])
                                    if 0 < pg <= len(src):
                                        pack_pdf.insert_pdf(src, from_page=pg-1, to_page=pg-1)
                                    src.close()
                                except Exception:
                                    pass

                    # 5. Hazard Directory & Access Points
                    pk_elr_label = swp_elr_from if swp_elr_from == swp_elr_to else f"{swp_elr_from} to {swp_elr_to}"
                    pk_mil_from = swp_dist_from1 or ''
                    pk_mil_to = swp_dist_to1 or ''
                    if not swp_hazards_df.empty:
                        hz_pk = swp_hazards_df.copy()
                        if 'Mileage  From' in hz_pk.columns:
                            hz_pk['Mileage  From'] = hz_pk['Mileage  From'].apply(decimal_to_miles_chains)
                        if 'Mileage To' in hz_pk.columns:
                            hz_pk['Mileage To'] = hz_pk['Mileage To'].apply(decimal_to_miles_chains)
                        dc = [c for c in ['ELR','ELR Name','Mileage  From','Mileage To','Hazard Description','Local Name','Track','Free Text'] if c in hz_pk.columns]
                        hz_buf = generate_pdf(hz_pk[dc].fillna(''), pk_elr_label, pk_mil_from, pk_mil_to, HAZARD_COLS, 'NWR Hazard Directory')
                        hz_src = _fitz.open(stream=hz_buf.read(), filetype="pdf")
                        pack_pdf.insert_pdf(hz_src)
                        hz_src.close()
                    if not swp_access_df.empty:
                        ap_pk = swp_access_df.copy()
                        if 'Mileage  From' in ap_pk.columns:
                            ap_pk['Mileage  From'] = ap_pk['Mileage  From'].apply(decimal_to_miles_chains)
                        dc = [c for c in ['ELR','ELR Name','Mileage  From','Hazard Description','Local Name','Track','Free Text','Google Maps'] if c in ap_pk.columns]
                        ap_buf = generate_pdf(ap_pk[dc].fillna(''), pk_elr_label, pk_mil_from, pk_mil_to, ACCESS_COLS, 'Access Points')
                        ap_src = _fitz.open(stream=ap_buf.read(), filetype="pdf")
                        pack_pdf.insert_pdf(ap_src)
                        ap_src.close()

                    pack_page_count = len(pack_pdf)
                    pack_bytes = pack_pdf.tobytes()
                    pack_pdf.close()
                    st.download_button(
                        f"⬇  Download Complete Pack ({pack_page_count} pages)",
                        data=pack_bytes,
                        file_name=f"Complete_SWP_Pack_{swp_elr_from}_{pack_ts}.pdf",
                        mime="application/pdf",
                        key="swp_pack_dl")

            if generate_btn:
                import io
                import tempfile
                import subprocess
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                from openpyxl.drawing.image import Image as XlImage
                from openpyxl.utils import get_column_letter
                import openpyxl.worksheet.pagebreak
                from datetime import datetime

                wb = Workbook()
                ws = wb.active
                ws.title = "Safe Work Pack"

                # Page setup — A4 portrait, fit to width
                ws.sheet_properties.pageSetUpPr.fitToPage = True
                ws.page_setup.fitToWidth = 1
                ws.page_setup.fitToHeight = 0
                ws.page_setup.paperSize = ws.PAPERSIZE_A4
                ws.page_margins.left = 0.4
                ws.page_margins.right = 0.4
                ws.page_margins.top = 0.3
                ws.page_margins.bottom = 0.3

                # Styles
                thin = Side(style='thin')
                med = Side(style='medium')
                brd = Border(left=thin, right=thin, top=thin, bottom=thin)
                brd_m = Border(left=med, right=med, top=med, bottom=med)
                grey = PatternFill('solid', fgColor='D9D9D9')
                white_f = PatternFill('solid', fgColor='FFFFFF')
                green_f = PatternFill('solid', fgColor='C6EFCE')
                nwr_f = PatternFill('solid', fgColor='003366')
                yellow_f = PatternFill('solid', fgColor='FFFF00')

                fn_b = Font(name='Arial', bold=True, size=10)
                fn_n = Font(name='Arial', size=10)
                fn_s = Font(name='Arial', size=8)
                fn_h = Font(name='Arial', bold=True, size=14)
                fn_t = Font(name='Arial', bold=True, size=18)
                fn_lb = Font(name='Arial', bold=True, size=9)
                fn_tk = Font(name='Arial', bold=True, size=12)
                fn_tg = Font(name='Arial', bold=True, size=11, color='008000')
                fn_sec = Font(name='Arial', bold=True, size=11, color='FFFFFF')
                fn_ft = Font(name='Arial', size=7, italic=True, color='666666')
                fn_red = Font(name='Arial', bold=True, size=9, color='CC2200')
                fn_big = Font(name='Arial', bold=True, size=12)
                fn_num = Font(name='Arial', bold=True, size=14)

                # Page 5 RT9909 — size 14 fonts for trackside readability
                fn_p5_label = Font(name='Arial', bold=True, size=14)
                fn_p5_val = Font(name='Arial', size=14)
                fn_p5_hdr = Font(name='Arial', bold=True, size=12)

                # Page 1 Cover — size 14 fonts
                fn_p1_label = Font(name='Arial', bold=True, size=14)
                fn_p1_val = Font(name='Arial', size=14)

                wt = Alignment(wrap_text=True, vertical='top')
                wc = Alignment(wrap_text=True, vertical='center')
                cc = Alignment(horizontal='center', vertical='center')
                cw = Alignment(horizontal='center', vertical='center', wrap_text=True)

                footer_text = "Issue 7 Dec 2025    Amended to Issue 13 019 Standard"
                prot_name = swp_prot_selected.split(' - ', 1)[1] if ' - ' in swp_prot_selected else swp_prot_selected

                # Column widths — 25 columns (A-Y)
                for c in range(1, 26):
                    ws.column_dimensions[get_column_letter(c)].width = 5.5

                # Helper functions
                def SR(r1, r2, c1, c2, font=None, fill=None, alignment=None, border=None):
                    for r in range(r1, r2+1):
                        for c in range(c1, c2+1):
                            cl = ws.cell(row=r, column=c)
                            if font: cl.font = font
                            if fill: cl.fill = fill
                            if alignment: cl.alignment = alignment
                            if border: cl.border = border

                def MC(r, c1, c2, val, font=fn_n, fill=None, align=None, h=None, r2=None):
                    """Merged cell helper. r2=end row for multi-row merge."""
                    er = r2 if r2 else r
                    if c2 > c1 or er > r:
                        ws.merge_cells(start_row=r, start_column=c1, end_row=er, end_column=c2)
                    cl = ws.cell(row=r, column=c1, value=val)
                    cl.font = font
                    if fill: cl.fill = fill
                    if align: cl.alignment = align
                    SR(r, er, c1, c2, border=brd)
                    if h: ws.row_dimensions[r].height = h
                    return er + 1

                def footer(r):
                    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
                    ws.cell(row=r, column=1, value="Issue 7 Dec 2025").font = fn_ft
                    ws.merge_cells(start_row=r, start_column=9, end_row=r, end_column=18)
                    ws.cell(row=r, column=9, value="Amended to Issue 13 019 Standard").font = fn_ft
                    ws.merge_cells(start_row=r, start_column=19, end_row=r, end_column=25)
                    ws.cell(row=r, column=19, value=swp_ref).font = fn_ft
                    ws.cell(row=r, column=19).alignment = Alignment(horizontal='right')
                    return r + 1

                def PB(r):
                    ws.row_breaks.append(openpyxl.worksheet.pagebreak.Break(id=r-1))
                    return r

                # ═══════════════════════════════════════
                # PAGE 1 — COVER
                # ═══════════════════════════════════════
                row = 1

                # Logo — top right
                try:
                    app_dir = os.path.dirname(os.path.abspath(__file__))
                    for lp in [os.path.join(app_dir, 'PPS_rail_logo.jpg'),
                               os.path.join(app_dir, 'PPS_Rail_Logo_with_Trademark.png'),
                               os.path.join(app_dir, 'PPS-logo-ol.png'),
                               os.path.join(app_dir, 'data', 'PPS_rail_logo.jpg')]:
                        if os.path.exists(lp):
                            img = XlImage(lp)
                            img.width = 220; img.height = 60
                            ws.add_image(img, 'S1')
                            break
                except Exception:
                    pass

                ws.row_dimensions[1].height = 32
                row = 2
                # Title — no border, just centred text
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=25)
                ws.cell(row=row, column=1, value="SAFE WORK PACK").font = fn_t
                ws.cell(row=row, column=1).alignment = cc
                ws.row_dimensions[row].height = 34
                row += 1
                row += 1  # blank row

                # Job details — FULL WIDTH col 1-25
                job = [
                    ("PRODUCED BY PLANNER", swp_planner_name, "Contact No", swp_planner_number),
                    ("DATE PRODUCED", datetime.now().strftime('%d/%m/%Y'), None, None),
                    ("LOCATION OF WORKS", elr_location, None, None),
                    ("NATURE OF WORKS", swp_nature_of_work, None, None),
                    ("METHOD OF PROTECTION", prot_name, None, None),
                    ("SWP REFERENCE No.", swp_ref, None, None),
                    ("RESPONSIBLE MANAGER", swp_rm_name, "Contact No", swp_rm_number),
                    ("PERSON IN CHARGE (Planning)", swp_coss_name, "Contact No", swp_coss_number),
                    ("ITEM No  &  WORKSITE Ref", f"{swp_item_no} - {swp_worksite}", None, None),
                    ("WEEK NO & DATE(S) OF WORKS", f"WK{swp_week} : {swp_from_date} - {swp_to_date}", None, None),
                ]
                for label, val, l2, v2 in job:
                    MC(row, 1, 8, f"  {label}", fn_p1_label, grey, wc, 24)
                    if l2:
                        MC(row, 9, 16, val, fn_p1_val, None, wt)
                        MC(row, 17, 19, l2, fn_p1_label, grey, wc)
                        MC(row, 20, 25, v2, fn_p1_val)
                        SR(row, row, 1, 25, border=brd)
                    else:
                        MC(row, 9, 25, val, fn_p1_val, None, wt)
                        SR(row, row, 1, 25, border=brd)
                    row += 1

                # Shift Contact Numbers — FULL WIDTH
                row = MC(row, 1, 25, "SHIFT CONTACT NUMBERS", fn_p1_label, grey, cc, 24)

                for lbl, c1, c2 in [("Name",1,7),("Duty",8,11),("Phone Number",12,18),("Shift Times",19,25)]:
                    MC(row, c1, c2, lbl, fn_p1_label, grey, cc)
                row += 1

                all_sc = list(shift_contacts) + [{'Name':'','Duty':'','Phone':'','Start':'','End':''}] * max(0, 5 - len(shift_contacts))
                for sc in all_sc[:5]:
                    MC(row, 1, 7, sc.get('Name',''), fn_p1_val)
                    MC(row, 8, 11, sc.get('Duty',''), fn_p1_val, None, cc)
                    MC(row, 12, 18, sc.get('Phone',''), fn_p1_val, None, cc)
                    MC(row, 19, 22, sc.get('Start',''), fn_p1_val, None, cc)
                    MC(row, 23, 25, sc.get('End',''), fn_p1_val, None, cc)
                    SR(row, row, 1, 25, border=brd)
                    row += 1

                row += 1

                # Yellow Planner Comments Box — only shows if comments exist
                if swp_planner_comments and swp_planner_comments.strip():
                    MC(row, 1, 25, swp_planner_comments, fn_big, yellow_f, cw, None, r2=row+2)
                    SR(row, row+2, 1, 25, border=brd_m)
                    row += 3
                    row += 1

                # Change authority — FULL WIDTH
                MC(row, 1, 10, "Reason and authority for change from\nplanned safe system of work", fn_s, grey, wc, 36, r2=row+1)
                MC(row, 11, 25, "", fn_n, None, None, None, r2=row+1)
                SR(row, row+1, 1, 25, border=brd_m)
                row += 2

                row += 1
                MC(row, 1, 10, "Name of Responsible Manager\nauthorising the change.", fn_s, grey, wc, 40, r2=row+1)
                MC(row, 11, 16, "", fn_n, None, None, None, r2=row+1)
                MC(row, 17, 20, "Signature/\nAuthority no", fn_s, grey, wc, None, r2=row+1)
                MC(row, 21, 25, "", fn_n, None, None, None, r2=row+1)
                SR(row, row+1, 1, 25, border=brd_m)
                row += 2

                row += 1
                # COSS to Planner Feedback — FULL WIDTH
                MC(row, 1, 8, "Coss To Planner\nFeedback", fn_big, None, wc, 60, r2=row+2)
                MC(row, 9, 25, "", fn_n, None, None, None, r2=row+2)
                SR(row, row+2, 1, 25, border=brd_m)
                row += 3

                row += 1
                MC(row, 1, 25, "THIS PACK MUST BE RETURNED TO THE Supervisor ON COMPLETION OF THE WORKS", Font(name='Arial', bold=True, size=9), None, cc, 20)
                row += 1
                MC(row, 1, 12, "Date Reviewed By planner", fn_b, grey, cc, 28, r2=row+1)
                MC(row, 13, 14, "", fn_n, None, None, None, r2=row+1)
                MC(row, 15, 20, "Planner Signature", fn_b, grey, cc, None, r2=row+1)
                MC(row, 21, 25, "", fn_n, None, None, None, r2=row+1)
                SR(row, row+1, 1, 25, border=brd_m)
                row += 2

                row += 1
                row = footer(row)
                row += 1
                row = PB(row)

                # ═══════════════════════════════════════
                # PAGE 2 — VALIDATION
                # ═══════════════════════════════════════
                row = MC(row, 1, 25, "SWP Validation Form", fn_h, None, Alignment(horizontal='left'), 24)

                # Rejected / Errors — full width 1-25
                MC(row, 1, 8, "Rejected", fn_b, grey, cc, 22)
                MC(row, 9, 16, "YES", fn_b, None, cc)
                MC(row, 17, 25, "NO", fn_b, None, cc)
                SR(row, row, 1, 25, border=brd)
                row += 1
                MC(row, 1, 8, "Errors / Changes", fn_b, grey, cc, 22)
                MC(row, 9, 16, "YES", fn_b, None, cc)
                MC(row, 17, 25, "NO", fn_b, None, cc)
                SR(row, row, 1, 25, border=brd)
                row += 2

                # Cyclical / Non-Cyclical / Repeat — full width 1-25
                for i, (lbl, c1, c2) in enumerate([("Cyclical",1,8),("Non-Cyclical",9,17),("Repeat",18,25)]):
                    is_sel = (swp_type == lbl)
                    MC(row, c1, c2, lbl, fn_b, green_f if is_sel else None, cc, 22)
                    SR(row, row, c1, c2, border=brd)
                row += 1

                # SWP Ref / Expiry / Date — row 1 (dates)
                MC(row, 1, 4, "SWP Ref.", fn_b, grey, cc, 22)
                MC(row, 5, 9, swp_ref, fn_n, None, cc)
                MC(row, 10, 14, "SWP expiry date", fn_lb, grey, cc)
                MC(row, 15, 18, swp_to_date, fn_n, None, cc)
                MC(row, 19, 21, "Date of Work", fn_lb, grey, cc)
                MC(row, 22, 25, f"{swp_from_date} - {swp_to_date}", fn_n, None, cw)
                SR(row, row, 1, 25, border=brd)
                row += 1
                # SWP Ref / Expiry / Date — row 2 (times)
                MC(row, 1, 18, "", fn_n, None, None, 22)
                MC(row, 19, 21, "Time of Work", fn_lb, grey, cc)
                MC(row, 22, 25, f"{swp_from_time} - {swp_to_time}", fn_n, None, cc)
                SR(row, row, 1, 25, border=brd)
                row += 1

                MC(row, 1, 8, "Brief Description of Work", fn_b, None, None, 20, r2=row+1)
                MC(row, 9, 25, swp_nature_of_work, fn_n, None, wt, None, r2=row+1)
                SR(row, row+1, 1, 25, border=brd)
                row += 2

                row += 1
                row = MC(row, 1, 25, "CREATED by: Planner", fn_b, grey, None, 20)
                row = MC(row, 1, 25, "I confirm this SWP has been checked and compliant with NR/L2/OHS/019, Appendix A and Form B", fn_s, None, wt, 20)

                MC(row, 1, 10, f"Planner Name: {swp_planner_name}", fn_n)
                MC(row, 11, 17, "Signature:", fn_lb)
                MC(row, 18, 25, "Date Issued:", fn_lb)
                SR(row, row, 1, 25, border=brd)
                row += 1

                row = MC(row, 1, 25, "VERIFIED by: Person in charge", fn_b, grey, None, 20)
                row = MC(row, 1, 25, "I confirm the following are appropriate for the task and included in the SWP (for guidance use Appendix C checklist). Circle Yes or No for each question, and sign the declaration below", fn_s, None, wt, 30)

                yn_pic = [
                    "The appropriate hierarchy of Safe System of Work has been selected",
                    "Task risk and any specific controls are suitable and sufficient",
                    "Necessary competence within team to undertake task",
                    "Any additional specific controls identified",
                    "Any necessary permit to work arrangements identified",
                    "The welfare facilities have been identified and are appropriate",
                ]
                for q in yn_pic:
                    MC(row, 1, 18, q, fn_n, None, wt)
                    MC(row, 19, 21, "Y", fn_tg, None, cc)
                    MC(row, 22, 25, "N", fn_n, None, cc)
                    SR(row, row, 1, 25, border=brd)
                    row += 1

                row = MC(row, 1, 25, "If any of the above statements are answered NO, reject the SWP and return it to the Planner.", fn_s, None, wt, 18)
                row = MC(row, 1, 25, "Comments if SWP rejected:", fn_s, None, None, 18)
                row += 1

                # PIC signature — proper merged boxes
                MC(row, 1, 10, "Name of Person in charge", fn_lb, grey)
                MC(row, 11, 17, "Signature:", fn_lb, grey)
                MC(row, 18, 25, "Date:", fn_lb, grey)
                SR(row, row, 1, 25, border=brd)
                row += 1
                MC(row, 1, 10, swp_coss_name, fn_n, None, None, 30, r2=row+1)
                MC(row, 11, 17, "", fn_n, None, None, None, r2=row+1)
                MC(row, 18, 25, "", fn_n, None, None, None, r2=row+1)
                SR(row, row+1, 1, 25, border=brd_m)
                row += 2

                # AUTHORISED by: RM
                row = MC(row, 1, 25, "AUTHORISED by: Responsible Manager", fn_b, grey, None, 20)
                row = MC(row, 1, 25, "Complete as part of review/discussion with person in charge. Circle Yes or No for each question and sign the declaration below.", fn_s, None, wt, 26)

                yn_rm = [
                    "Work content is understood by the person in charge",
                    "Necessary competence within team to undertake task",
                    "Task risk and any specific controls are suitable and sufficient",
                    "The appropriate hierarchy of Safe System of Work has been selected",
                    "Any additional specific controls identified",
                    "The welfare facilities have been identified and are appropriate",
                ]
                for q in yn_rm:
                    MC(row, 1, 18, q, fn_n, None, wt)
                    MC(row, 19, 21, "Y", fn_tg, None, cc)
                    MC(row, 22, 25, "N", fn_n, None, cc)
                    SR(row, row, 1, 25, border=brd)
                    row += 1

                row = MC(row, 1, 25, "Responsible Manager's authorisation and confirmation this SWP is complete, and includes any specific additional information required to manage risk on site (cannot be the same person as the verifier). If any of the above statements are answered NO, reject the SWP.", fn_s, None, wt, 30)

                # RM signature — proper merged boxes with height
                MC(row, 1, 10, "Print Name:", fn_lb, grey)
                MC(row, 11, 17, "Signature or Authority Number:", fn_lb, grey)
                MC(row, 18, 25, "Date:", fn_lb, grey)
                SR(row, row, 1, 25, border=brd)
                row += 1
                MC(row, 1, 10, swp_rm_name, fn_n, None, None, 36, r2=row+1)
                MC(row, 11, 17, "", fn_n, None, None, None, r2=row+1)
                MC(row, 18, 25, "", fn_n, None, None, None, r2=row+1)
                SR(row, row+1, 1, 25, border=brd_m)
                row += 2

                # ACCEPTED by PIC on site
                row = MC(row, 1, 25, "ACCEPTED by: Person in charge on site", fn_b, grey, None, 20)
                row = MC(row, 1, 25, "Person in charge, at the site of work completes this section. Endorse declaration below.", fn_s, None, wt, 20)
                row = MC(row, 1, 25, "I Accept / Reject this SWP (Please circle)", fn_b, None, None, 20)
                row = MC(row, 1, 25, "If rejected, detail briefly change(s) needed:\n\nGeneral comments:", fn_s, None, wt, 44)

                MC(row, 1, 13, "Person in charge (Name):", fn_lb, grey)
                MC(row, 14, 25, "Signature:", fn_lb, grey)
                SR(row, row, 1, 25, border=brd)
                row += 1
                MC(row, 1, 13, "", fn_n, None, None, 30, r2=row+1)
                MC(row, 14, 25, "", fn_n, None, None, None, r2=row+1)
                SR(row, row+1, 1, 25, border=brd_m)
                row += 2
                MC(row, 1, 13, "Date:", fn_lb, grey)
                MC(row, 14, 25, "Location i.e. Site:", fn_lb, grey)
                SR(row, row, 1, 25, border=brd)
                row += 1
                MC(row, 1, 13, "", fn_n, None, None, 24, r2=row)
                MC(row, 14, 25, "", fn_n, None, None, None, r2=row)
                SR(row, row, 1, 25, border=brd_m)
                row += 1

                row = footer(row)
                row += 1
                row = PB(row)

                # ═══════════════════════════════════════
                # PAGE 3 — RISKS & RUNAWAY
                # ═══════════════════════════════════════
                MC(row, 1, 7, "SWP Ref.", fn_b, grey, None, 20, r2=row+1)
                # Force right border on column 7
                for r_tmp in [row, row+1]:
                    ws.cell(row=r_tmp, column=7).border = Border(right=thin, top=thin if r_tmp == row else Side(style=None), bottom=thin if r_tmp == row+1 else Side(style=None))
                MC(row, 8, 12, swp_ref, fn_n, None, cc, None, r2=row+1)
                # Clean value area - outer border only
                for r_tmp in [row, row+1]:
                    for c_tmp in range(8, 13):
                        cell = ws.cell(row=r_tmp, column=c_tmp)
                        l = thin if c_tmp == 8 else Side(style=None)
                        ri = thin if c_tmp == 12 else Side(style=None)
                        t = thin if r_tmp == row else Side(style=None)
                        b = thin if r_tmp == row+1 else Side(style=None)
                        cell.border = Border(left=l, right=ri, top=t, bottom=b)
                MC(row, 13, 16, "Date & Time of", fn_lb, grey)
                MC(row, 17, 25, f"{swp_from_date} - {swp_to_date}", fn_n, None, wt)
                SR(row, row, 13, 25, border=brd)
                row += 1
                MC(row, 13, 16, "Work", fn_lb, grey)
                MC(row, 17, 25, f"{swp_from_time} - {swp_to_time}", fn_n)
                SR(row, row, 13, 25, border=brd)
                row += 1

                MC(row, 1, 7, "Brief Description of work", fn_b, grey, None, 20, r2=row+1)
                MC(row, 8, 25, swp_nature_of_work, fn_n, None, wt, None, r2=row+1)
                SR(row, row+1, 1, 25, border=brd)
                row += 2

                row += 1
                row = MC(row, 1, 25, "Risk Identified and controls to be applied (Not applicable where task briefing sheets are included)", fn_b, grey, wt, 20)

                # Risk table — 5 columns
                for lbl, c1, c2 in [("Specific Risk Requiring Control",1,8),("Control to be applied",9,13),("Person in charge /RM initials",14,16),("Control Delegated to: (name)",17,19),("Acceptance of Control by:",20,25)]:
                    MC(row, c1, c2, lbl, fn_lb, grey, cw, None, r2=row+1)
                row += 2
                for _ in range(4):
                    for c1, c2 in [(1,8),(9,13),(14,16),(17,19),(20,25)]:
                        MC(row, c1, c2, "", fn_n, None, None, None, r2=row+1)
                    row += 2

                row += 1
                row = MC(row, 1, 25, "Runaway Risk analysis", fn_b, grey, None, 20)

                for lbl, c1, c2 in [("Question",1,13),("Answer",14,16),("Comment",18,25)]:
                    MC(row, c1, c2, lbl, fn_b, grey, cc)
                SR(row, row, 1, 25, border=brd)
                row += 1

                rq = [
                    "Are the works taking place on or near the line?",
                    "Could your work potentially Lead to a Runaway? I.e. involve the use of equipment subject to Runaway and control requirements (Trolleys, Trailers, Manually propelled rail handling equipment)",
                    "Are my works within a Possession or adjacent to a Possession?",
                    "Are my works on a gradient steeper than 1 in 100 or is there a gradient within 5 miles of my works?",
                    "Is the site of work at risk of a runaway from a 3rd Party",
                ]
                for i, q in enumerate(rq):
                    h = 22 if len(q) < 60 else 40
                    MC(row, 1, 13, q, fn_n, None, wt, h)
                    ans = runaway_answers[i] if i < len(runaway_answers) else "No"
                    MC(row, 14, 16, ans, fn_tg if ans == "Yes" else fn_n, None, cc)
                    MC(row, 18, 25, runaway_comments[i] if i < len(runaway_comments) else "", fn_n, None, wt)
                    SR(row, row, 1, 25, border=brd)
                    row += 1

                row = MC(row, 1, 25, "Responsible Manager and PIC or Site Manager review of Runaway Risk Review", fn_b, grey, wt, 20)

                # RM - labels row
                MC(row, 1, 13, "RM Name:", fn_lb, grey)
                MC(row, 14, 25, "Signature:", fn_lb, grey)
                SR(row, row, 1, 25, border=brd)
                row += 1
                # RM - value row
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=13)
                ws.cell(row=row, column=1, value=swp_rm_name).font = fn_n
                ws.merge_cells(start_row=row, start_column=14, end_row=row, end_column=25)
                ws.row_dimensions[row].height = 36
                for c_tmp in range(1, 26):
                    cell = ws.cell(row=row, column=c_tmp)
                    l = med if c_tmp == 1 else (med if c_tmp == 14 else Side(style=None))
                    ri = med if c_tmp == 25 else (med if c_tmp == 13 else Side(style=None))
                    cell.border = Border(left=l, right=ri, top=Side(style=None), bottom=med)
                row += 1

                # PIC - labels row
                MC(row, 1, 13, "PIC Name:", fn_lb, grey)
                MC(row, 14, 25, "Signature:", fn_lb, grey)
                SR(row, row, 1, 25, border=brd)
                row += 1
                # PIC - value row
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=13)
                ws.cell(row=row, column=1, value=swp_coss_name).font = fn_n
                ws.merge_cells(start_row=row, start_column=14, end_row=row, end_column=25)
                ws.row_dimensions[row].height = 36
                for c_tmp in range(1, 26):
                    cell = ws.cell(row=row, column=c_tmp)
                    l = med if c_tmp == 1 else (med if c_tmp == 14 else Side(style=None))
                    ri = med if c_tmp == 25 else (med if c_tmp == 13 else Side(style=None))
                    cell.border = Border(left=l, right=ri, top=Side(style=None), bottom=med)
                row += 1

                row = footer(row)
                row += 1
                row = PB(row)

                # ═══════════════════════════════════════
                # PAGE 4 — FORM B
                # ═══════════════════════════════════════
                row = MC(row, 1, 25, "Form B", fn_h, None, None, 24)

                MC(row, 1, 5, "Date(s) of Work:", fn_b, grey)
                MC(row, 6, 12, f"{swp_from_date}-{swp_to_date}", fn_n)
                MC(row, 13, 15, "Time of Work:", fn_b, grey)
                MC(row, 16, 25, f"{swp_from_time} - {swp_to_time}", fn_n)
                SR(row, row, 1, 25, border=brd)
                row += 1

                MC(row, 1, 5, "Location", fn_b, grey, None, None, r2=row+1)
                MC(row, 6, 12, elr_location, fn_n, None, wt, None, r2=row+1)
                MC(row, 13, 15, "SWP Ref No:", fn_b, grey, None, None, r2=row+1)
                MC(row, 16, 25, swp_ref, fn_n, None, wt, None, r2=row+1)
                SR(row, row+1, 1, 25, border=brd)
                row += 2

                row += 1
                row = MC(row, 1, 25, "The following safe systems of work are organised in priority order. Each should be considered with the highest achievable system consistent with the nature, location and duration of the work to be selected. Following selection of the SSOW the Responsible Manager must sign the 'Form B' authorisation", fn_s, None, wt, 36)

                row = MC(row, 1, 25, "If a safe system of work IS selected, tick the YES box next to the system.", fn_n, None, None, 18)
                row += 1
                row = MC(row, 1, 25, "S.S.O.W: 1 to 3.", fn_b, None, None, 20)

                row = MC(row, 1, 25, "If S.S.O.W protection between 1 & 3 IS NOT selected, tick the NO box next to the method AND provide an explanation in the box below - then consider the next method. If none of the methods of protection can be selected proceed to warning systems (methods 4 to 6)", fn_s, None, wt, 30)

                row += 1
                MC(row, 1, 21, "          Hierarchy of control for operational risks (Protection)", fn_b)
                MC(row, 22, 25, "SELECTED", fn_b, None, cc)
                SR(row, row, 1, 25, border=brd)
                row += 1
                MC(row, 22, 22, "YES", fn_lb, grey, cc)
                MC(row, 23, 25, "NO", fn_lb, grey, cc)
                SR(row, row, 22, 25, border=brd)
                row += 1

                # Methods 1-3 with YES/NO ticks
                prot_methods_fb = [
                    (1, "Safeguarded", sg_reason if prot_num >= 2 else ""),
                    (2, "Fenced", fence_reason if prot_num >= 3 else ""),
                    (3, "Separated", ""),
                ]
                for num, name, reason in prot_methods_fb:
                    is_sel = prot_num == num
                    MC(row, 1, 3, str(num), fn_num, None, cc, None, r2=row+2)
                    MC(row, 4, 21, f"{name} If this method is NOT selected, please give reasons here.", fn_n, None, wt)
                    # YES column — tick if selected
                    MC(row, 22, 22, "\u2713" if is_sel else "", fn_tk, green_f if is_sel else None, cc, None, r2=row+2)
                    # NO column — tick if NOT selected
                    MC(row, 23, 25, "\u2713" if not is_sel else "", fn_tk, None, cc, None, r2=row+2)
                    SR(row, row+2, 1, 25, border=brd)
                    row += 1
                    MC(row, 4, 21, reason, fn_n, None, wt, None, r2=row+1)
                    row += 2

                row += 1
                row = MC(row, 1, 25, "Warning Systems 4 to 6", fn_b, None, None, 20)

                row = MC(row, 1, 25, "Supplementary questions A, B; C & D must be answered before using warning systems 4 to 6. If the answer to any of the supplementary questions is YES - THEN WORK MUST BE PLANNED USING METHODS 1 TO 3. If all of the questions are answered NO, continue to methods 4 to 6.", fn_s, None, wt, 36)

                row += 1
                MC(row, 1, 21, "           SUPPLEMENTARY QUESTIONS for  4 to 6", fn_b)
                MC(row, 22, 22, "YES", fn_lb, grey, cc)
                MC(row, 23, 25, "NO", fn_lb, grey, cc)
                SR(row, row, 1, 25, border=brd)
                row += 1

                supp = [
                    ("A", "Is the line speed greater than 125mph (200kph)? (Answer NO if a temporary or emergency speed restriction to 125mph (200kph) or less applies)?"),
                    ("B", "Does the total warning time required exceed 45 seconds?"),
                    ("C", "Are there three or more lines open to traffic between the site of work and the designated position(s) of safety?"),
                    ("D", "Does the Network Rail Hazard Directory prohibit 4 to 6 working at this location?"),
                ]
                for letter, q in supp:
                    MC(row, 1, 2, letter, fn_b, None, cc, None, r2=row+1)
                    MC(row, 3, 21, q, fn_n, None, wt, None, r2=row+1)
                    MC(row, 22, 22, "", fn_n, None, cc, None, r2=row+1)
                    MC(row, 23, 25, "", fn_n, None, cc, None, r2=row+1)
                    SR(row, row+1, 1, 25, border=brd)
                    row += 2

                row += 1
                # Warning 4-6
                MC(row, 1, 21, "          Hierarchy of control for operational risks (Warning)", fn_b)
                MC(row, 22, 25, "SELECTED", fn_b, None, cc)
                SR(row, row, 1, 25, border=brd)
                row += 1
                MC(row, 22, 22, "YES", fn_lb, grey, cc)
                MC(row, 23, 25, "NO", fn_lb, grey, cc)
                SR(row, row, 22, 25, border=brd)
                row += 1

                warn = [
                    (4, "Warning systems permanent-train activated equipment"),
                    (5, "Warning systems portable - train activated equipment"),
                    (6, "Lookout warning \u2013 maximum permissible line or temporarily restricted to 25 mph"),
                ]
                for num, name in warn:
                    is_sel = prot_num == num
                    MC(row, 1, 3, str(num), fn_num, None, cc, None, r2=row+2)
                    reason_text = f"{name} If this method is NOT selected, please give reasons here." if num < 6 else name
                    MC(row, 4, 21, reason_text, fn_n, None, wt, None, r2=row+2)
                    MC(row, 22, 22, "\u2713" if is_sel else "", fn_tk, green_f if is_sel else None, cc, None, r2=row+2)
                    MC(row, 23, 25, "\u2713" if not is_sel else "", fn_tk, None, cc, None, r2=row+2)
                    SR(row, row+2, 1, 25, border=brd)
                    row += 3

                row = footer(row)
                row += 1
                row = PB(row)

                # ═══════════════════════════════════════
                # PAGE 5/6 — RT9909 + SSOW (combined if fits)
                # ═══════════════════════════════════════
                row = MC(row, 1, 25, "RECORD OF ARRANGEMENTS AND BRIEFING FORM RT9909", fn_h, None, None, 24)
                row = MC(row, 1, 25, "GENERAL INFORMATION *where the work is pre-planned, these parts of the form should be completed before it is provided to the COSS/IWA.", fn_s, None, wt, 24)

                # COSS / Sentinel — labels row
                MC(row, 1, 10, "Name of COSS/IWA", fn_p5_label, grey)
                MC(row, 11, 25, "Sentinel Card No.", fn_p5_label, grey)
                SR(row, row, 1, 25, border=brd)
                row += 1
                # Values row
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
                ws.cell(row=row, column=1, value=swp_coss_name).font = fn_p5_val
                ws.merge_cells(start_row=row, start_column=11, end_row=row, end_column=25)
                ws.row_dimensions[row].height = 30
                for c_tmp in range(1, 26):
                    cell = ws.cell(row=row, column=c_tmp)
                    l = med if c_tmp == 1 else (med if c_tmp == 11 else Side(style=None))
                    ri = med if c_tmp == 25 else (med if c_tmp == 10 else Side(style=None))
                    cell.border = Border(left=l, right=ri, top=Side(style=None), bottom=med)
                row += 1

                MC(row, 1, 5, "Date", fn_p5_label, grey)
                MC(row, 6, 25, "", fn_p5_val)
                SR(row, row, 1, 25, border=brd)
                row += 1

                MC(row, 1, 5, "Nature of Work*", fn_p5_label, grey, wt)
                MC(row, 6, 25, swp_nature_of_work, fn_p5_val, None, wt)
                SR(row, row, 1, 25, border=brd)
                row += 1

                MC(row, 1, 5, "Time Work Started", fn_p5_label, grey, wt, 38)
                MC(row, 6, 9, "", fn_p5_val)
                MC(row, 10, 12, "Time Work Finished", fn_p5_label, grey, wt)
                MC(row, 13, 25, "", fn_p5_val)
                SR(row, row, 1, 25, border=brd)
                row += 1

                # Location and lines
                lines_affected = elr_location
                if swp_line_data:
                    lines_affected += " / " + ", ".join([ld['Line Name'] for ld in swp_line_data[:2]])
                MC(row, 1, 5, "Location and Lines\nAffected*", fn_p5_label, grey, wt, 40)
                MC(row, 6, 25, lines_affected, fn_p5_val, None, wt)
                SR(row, row, 1, 25, border=brd)
                row += 1

                # Signaller
                sb_text = ""
                ecr_text = ""
                if swp_boxes:
                    box = swp_boxes[0]
                    sb_text = f"{box['Signal Box']}: Tel {box['Phone']}" if box['Phone'] else box['Signal Box']
                    if len(swp_boxes) > 1:
                        b2 = swp_boxes[1]
                        sb_text += f"  {b2['Signal Box']}: Tel {b2['Phone']}" if b2['Phone'] else f"  {b2['Signal Box']}"
                    if box.get('ECO'):
                        ecr_text = box['ECO']
                        if box.get('ECO Phone'): ecr_text += f": Tel {box['ECO Phone']}"

                MC(row, 1, 5, "How to contact the\nSignaller* in an\nemergency", fn_p5_label, grey, wt, 56)
                MC(row, 6, 25, sb_text, fn_p5_val, None, wt)
                SR(row, row, 1, 25, border=brd)
                row += 1

                MC(row, 1, 5, "Phone Number of\nElectrical Control\nRoom", fn_p5_label, grey, wt, 56)
                MC(row, 6, 25, ecr_text, fn_p5_val, None, wt)
                SR(row, row, 1, 25, border=brd)
                row += 1

                # Lines at site table
                MC(row, 1, 5, "Lines at the Site*", fn_p5_label, grey, wt)
                MC(row, 6, 11, "Direction (any SLW etc?)", fn_p5_hdr, grey)
                MC(row, 12, 17, "Open or Blocked*", fn_p5_hdr, grey)
                MC(row, 18, 25, "Speed (Line or T/ESR)", fn_p5_hdr, grey)
                SR(row, row, 1, 25, border=brd)
                row += 1

                for ld in swp_line_data:
                    MC(row, 1, 5, ld['Abbreviation'], fn_p5_val)
                    MC(row, 6, 11, "", fn_p5_val)
                    MC(row, 12, 17, ld['Status'], fn_p5_val)
                    MC(row, 18, 25, "", fn_p5_val)
                    SR(row, row, 1, 25, border=brd)
                    row += 1

                for _ in range(max(0, 3 - len(swp_line_data))):
                    for c1, c2 in [(1,5),(6,11),(12,17),(18,25)]:
                        MC(row, c1, c2, "", fn_p5_val)
                    SR(row, row, 1, 25, border=brd)
                    row += 1

                # RT9909 detail fields
                rt_fields = [
                    ("Task Key Risks and Controls*\n(Include Risks from Other parties)", swp_task_risks),
                    ("Permits Required ( Lifting Plans , Electrical ,\nIsolation ,Hot works ,Confined Spaces)", swp_permits),
                    ("Welfare Arrangements and their location*", "Local Welfare Facilities / As Per Task Brief"),
                ]
                fa_text = "First Aider on Site:"
                if swp_nearest_ae:
                    h = swp_nearest_ae[0]
                    fa_text += f"\nNearest A&E: {h['Hospital']}, {h['Address']}, {h['Postcode']}"
                rt_fields.append(("First Aid Arrangements*", fa_text))
                access_text = ", ".join(swp_access_selected) if swp_access_selected else ""
                if swp_access_manual:
                    access_text += ("\n" if access_text else "") + swp_access_manual
                rt_fields.append(("Access & Egress Arrangements to/from\nworking area *", access_text))
                rt_fields.append(("Hazards associated with access/egress\n(conductor rails, tripping, vegetation,\noverhead cables or OLE, etc.)*", swp_access_hazards))
                rt_fields.append(("Hazards associated with the site\n(conductor rails, tripping, vegetation,\noverhead cables or OLE, etc.)*", swp_site_hazards))

                for label, val in rt_fields:
                    # Taller rows for size 14 fonts
                    h = 50 if '\n' in label else 38
                    MC(row, 1, 10, label, fn_p5_label, grey, wt, h, r2=row+1)
                    MC(row, 11, 25, val, fn_p5_val, None, wt, None, r2=row+1)
                    SR(row, row+1, 1, 25, border=brd)
                    row += 2

                # Limits — on Page 5 with RT9909
                MC(row, 1, 10, "Limits of the working area and how\nthese are defined*", fn_p5_label, grey, wt, 44, r2=row+1)
                MC(row, 11, 25, swp_limits, fn_p5_val, None, wt, None, r2=row+1)
                SR(row, row+1, 1, 25, border=brd_m)
                row += 2

                row = footer(row)
                row += 1
                row = PB(row)

                # ═══════════════════════════════════════
                # PAGE 7 — SSOW & DECLARATIONS
                # ═══════════════════════════════════════

                row += 1
                row = MC(row, 1, 25, "SAFE SYSTEM OF WORK", fn_b, grey, cc, 20)

                # SSOW tick table
                all_methods = [
                    "1  Safeguarded", "2  Fenced", "3  Separated",
                    "4  Warning System Permanent Train Operated",
                    "5  Warning System Portable Train Operated",
                    "6  Lookout Warning"
                ]
                walking_num = int(swp_ssow_walking[0]) if swp_ssow_walking else prot_num
                working_num = int(swp_ssow_working[0]) if swp_ssow_working else prot_num

                # Headers — match template layout
                MC(row, 1, 8, 'Tick the relevant box.  Only tick the \u201cPlanned\u201d column if you have been provided with a planned safe system of work', fn_n, grey, cw, 40, r2=row+1)
                MC(row, 9, 16, "Walking on or near the line to/from the working area", fn_n, None, cw, None, r2=row+1)
                MC(row, 17, 25, "Whilst carrying out the work", fn_n, None, cw, None, r2=row+1)
                SR(row, row+1, 9, 25, border=brd)
                ws.cell(row=row, column=9).alignment = cw
                ws.cell(row=row, column=17).alignment = cw
                # Clean instruction text box
                for r_tmp in [row, row+1]:
                    for c_tmp in range(1, 9):
                        cell = ws.cell(row=r_tmp, column=c_tmp)
                        l = thin if c_tmp == 1 else Side(style=None)
                        ri = thin if c_tmp == 8 else Side(style=None)
                        t = thin if r_tmp == row else Side(style=None)
                        b = thin if r_tmp == row+1 else Side(style=None)
                        cell.border = Border(left=l, right=ri, top=t, bottom=b)
                        cell.fill = grey
                row += 2

                # Planned / Actual sub-headers — equal widths
                MC(row, 1, 8, "", fn_n, grey)
                MC(row, 9, 12, "Planned*", fn_lb, grey, cc)
                MC(row, 13, 16, "Actual", fn_lb, grey, cc)
                MC(row, 17, 21, "Planned*", fn_lb, grey, cc)
                MC(row, 22, 25, "Actual", fn_lb, grey, cc)
                SR(row, row, 1, 25, border=brd)
                row += 1

                for m in all_methods:
                    num = int(m[0])
                    MC(row, 1, 8, m, fn_b if num == walking_num or num == working_num else fn_n)
                    MC(row, 9, 12, "\u2713" if num == walking_num else "", fn_tg, green_f if num == walking_num else None, cc)
                    MC(row, 13, 16, "", fn_n, None, cc)
                    MC(row, 17, 21, "\u2713" if num == working_num else "", fn_tg, green_f if num == working_num else None, cc)
                    MC(row, 22, 25, "", fn_n, None, cc)
                    SR(row, row, 1, 25, border=brd)
                    row += 1

                row += 1
                row = MC(row, 1, 25, "1 to 3 WORKING ONLY (complete as applicable)*", fn_b, grey, None, 20)

                for label, val in [
                    ("Type of Fence (fenced only)", swp_fence_type if prot_num == 2 else "N/A"),
                    ("Distance from the Line (fenced only)", swp_fence_dist if prot_num == 2 else "N/A"),
                    ("Separation distance (Site Warden only)", swp_sep_dist if prot_num == 3 else "N/A"),
                    ("How Site Warden will give the warning", swp_sep_warning if prot_num == 3 else "N/A"),
                ]:
                    MC(row, 1, 8, label, fn_n, grey)
                    MC(row, 9, 25, val, fn_n)
                    SR(row, row, 1, 25, border=brd)
                    row += 1

                row = MC(row, 1, 25, "4 to 6 WORKING ONLY", fn_b, grey, None, 20)
                MC(row, 1, 8, "How the warning will be given*", fn_n, grey)
                MC(row, 9, 25, "", fn_n)
                SR(row, row, 1, 25, border=brd)
                row += 1
                MC(row, 1, 8, "Location(s) of position(s) of Safety", fn_n, grey)
                MC(row, 9, 25, "", fn_n)
                SR(row, row, 1, 25, border=brd)
                row += 1

                # Wardens table
                row = MC(row, 1, 25, "Details of any use of Site Wardens, ATWS Operator or Lookout(s), First Aiders, Banksmen\n(TOWS, LOWS, Pee Wee, Distant, Intermediate, Site, Machine or Touch)", fn_s, grey, wt, 30)

                for lbl, c1, c2 in [("Name",1,8),("Sentinel Card No.",9,13),("Location/Position",14,18),("Role",19,25)]:
                    MC(row, c1, c2, lbl, fn_lb, grey, cc)
                SR(row, row, 1, 25, border=brd)
                row += 1
                for _ in range(3):
                    for c1, c2 in [(1,8),(9,13),(14,18),(19,25)]:
                        MC(row, c1, c2, "", fn_n)
                    SR(row, row, 1, 25, border=brd)
                    row += 1

                row += 1
                row = MC(row, 1, 25, "DECLARATION (Each member of the group to sign and confirm they have been briefed and understand the safe system of work arrangements to be implemented and the Site and Task risks briefed contained within this Safe Work Pack)", fn_s, grey, wt, 30)

                row += 1
                for lbl, c1, c2 in [("Name & Signature",1,8),("Sentinel Card No.",9,13),("Name & Signature",14,18),("Sentinel Card No.",19,25)]:
                    MC(row, c1, c2, lbl, fn_lb, grey, cc)
                SR(row, row, 1, 25, border=brd)
                row += 1
                for _ in range(8):
                    for c1, c2 in [(1,8),(9,13),(14,18),(19,25)]:
                        MC(row, c1, c2, "", fn_n)
                    SR(row, row, 1, 25, border=brd)
                    ws.row_dimensions[row].height = 28
                    row += 1

                row += 1
                # COSS Declaration
                MC(row, 1, 25, "COSS/IWA DECLARATION. I have made the above arrangements and am satisfied that all members of the work group understand the safe system of work.", fn_n, None, wt, None, r2=row+1)
                SR(row, row+1, 1, 25, border=brd)
                row += 2

                row += 1
                MC(row, 1, 10, "Name & Signature", fn_lb, grey, None, 30, r2=row+1)
                MC(row, 11, 25, "", fn_n, None, None, None, r2=row+1)
                SR(row, row+1, 1, 25, border=brd_m)
                row += 2

                MC(row, 1, 10, "COSS/IWA MUST identify how he/she has\nverified & confirmed his/her location", fn_lb, grey, wt, 30, r2=row+1)
                MC(row, 11, 25, "", fn_n, None, None, None, r2=row+1)
                SR(row, row+1, 1, 25, border=brd_m)
                row += 2

                row += 1
                MC(row, 1, 25, "I have relieved the above COSS/IWA and I am satisfied with the safe system. I have re-briefed the work group and am satisfied that all members of the work group understand the safe system of work.", fn_n, None, wt, None, r2=row+1)
                SR(row, row+1, 1, 25, border=brd)
                row += 2

                MC(row, 1, 10, "Name & Signature", fn_lb, grey, None, 30, r2=row+1)
                MC(row, 11, 13, "", fn_n, None, None, None, r2=row+1)
                MC(row, 14, 17, "Name & Signature", fn_lb, grey, None, None, r2=row+1)
                MC(row, 18, 25, "", fn_n, None, None, None, r2=row+1)
                SR(row, row+1, 1, 25, border=brd_m)
                row += 2

                row = footer(row)

                # ═══════════════════════════════════════
                # SAVE AND CONVERT TO PDF
                # ═══════════════════════════════════════
                safe_ref = re.sub(r'[^a-zA-Z0-9_-]', '_', swp_ref) if swp_ref else 'SWP'

                tmp_dir = tempfile.mkdtemp()
                xlsx_path = os.path.join(tmp_dir, f"SWP_{safe_ref}.xlsx")
                wb.save(xlsx_path)

                # Convert to PDF
                pdf_bytes = None
                try:
                    subprocess.run([
                        'libreoffice', '--headless', '--convert-to', 'pdf',
                        '--outdir', tmp_dir, xlsx_path
                    ], capture_output=True, timeout=60)
                    pdf_path = os.path.join(tmp_dir, f"SWP_{safe_ref}.pdf")
                    if os.path.exists(pdf_path):
                        with open(pdf_path, 'rb') as pf:
                            pdf_bytes = pf.read()
                except Exception as conv_err:
                    st.warning(f"PDF conversion issue: {conv_err}")

                with open(xlsx_path, 'rb') as xf:
                    xlsx_bytes = xf.read()

                # Store in session state for Complete Pack
                if pdf_bytes:
                    st.session_state['swp_pdf_bytes'] = pdf_bytes

                # PRIMARY: PDF
                if pdf_bytes:
                    st.download_button(
                        label="\U0001F4C4  Download SWP (PDF)",
                        data=pdf_bytes,
                        file_name=f"SWP_{safe_ref}.pdf",
                        mime="application/pdf",
                        key="swp_pdf_dl"
                    )

                # SECONDARY: Excel backup
                st.download_button(
                    label="\U0001F4DD  Download SWP (Excel \u2014 editable backup)",
                    data=xlsx_bytes,
                    file_name=f"SWP_{safe_ref}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="swp_excel_dl"
                )

                if pdf_bytes:
                    st.markdown(f"""
                    <div class="pps-card pps-card-green" style="margin-top:1rem">
                      <span class="badge badge-green">&#10003; SWP Generated</span>
                      <span style="margin-left:1rem;font-size:0.85rem">PDF + Excel ready. Ref: {swp_ref}</span>
                    </div>
                    """, unsafe_allow_html=True)
                else:
                    st.markdown(f"""
                    <div class="pps-card pps-card-amber" style="margin-top:1rem">
                      <span class="badge badge-amber">Excel Only</span>
                      <span style="margin-left:1rem;font-size:0.85rem">PDF conversion not available \u2014 download Excel and print to PDF. Ref: {swp_ref}</span>
                    </div>
                    """, unsafe_allow_html=True)

                try:
                    import shutil
                    shutil.rmtree(tmp_dir, ignore_errors=True)
                except Exception:
                    pass



        else:
            if not st.session_state.get('swp_data'):
                st.markdown(f"""
                <div style="font-size:0.85rem;color:{COLOURS['muted']};margin-top:2rem;text-align:center">
                  Paste a row from the Possession Tracker above and press <b>BUILD SWP</b> to get started.
                </div>
                """, unsafe_allow_html=True)

